import sys
import threading
import time
import struct
import re
import os
import logging
import serial
import serial.tools.list_ports
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from protocol_thread import ProtocolThread

# 抑制 pymodbus 调试日志输出
logging.getLogger('pymodbus').setLevel(logging.WARNING)

# ---------- 协议常量 ----------
FRAME_HEADER = 0xAA

CMD_START_DISCHARGE    = 0x01
CMD_STOP_DISCHARGE     = 0x02
CMD_START_CHARGE       = 0x03
CMD_STOP_CHARGE        = 0x04
CMD_SET_OVERC_PROT     = 0x05
CMD_SET_OVERD_PROT     = 0x06
CMD_SET_CYCLE_NUM      = 0x07
CMD_SET_CHARGE_VOLT    = 0x08
CMD_SET_CHARGE_CUR     = 0x09
CMD_SET_BAT_TYPE       = 0x0A
CMD_BUZZER_ON          = 0x0B
CMD_BUZZER_OFF         = 0x0C
CMD_BATCH_READ         = 0x10

CMD_ALL_START_CHARGE    = 0x20
CMD_ALL_STOP_CHARGE     = 0x21
CMD_ALL_START_DISCHARGE = 0x22
CMD_ALL_STOP_DISCHARGE  = 0x23
CMD_ALL_SET_CYCLE_NUM   = 0x24
CMD_ALL_SET_BAT_TYPE    = 0x25
CMD_ALL_EMERGENCY_STOP  = 0x26
CMD_ALL_BUZZER_ON       = 0x27
CMD_ALL_BUZZER_OFF      = 0x28

MODULE_COUNT = 64
MODULE_ADDRS = list(range(1, MODULE_COUNT + 1))  # Modbus 地址 1-64

CMD_NAMES = {
    CMD_START_DISCHARGE: "Start Discharge", CMD_STOP_DISCHARGE: "Stop Discharge",
    CMD_START_CHARGE: "Start Charge", CMD_STOP_CHARGE: "Stop Charge",
    CMD_SET_OVERC_PROT: "Set Overcharge Protect", CMD_SET_OVERD_PROT: "Set Overdischarge Protect",
    CMD_SET_CYCLE_NUM: "Set Cycle Number", CMD_SET_CHARGE_VOLT: "Set Charge Voltage",
    CMD_SET_CHARGE_CUR: "Set Charge Current", CMD_SET_BAT_TYPE: "Set Battery Type",
    CMD_BATCH_READ: "Batch Read",
    CMD_BUZZER_ON: "Buzzer On", CMD_BUZZER_OFF: "Buzzer Off",
    CMD_ALL_START_CHARGE: "All Start Charge", CMD_ALL_STOP_CHARGE: "All Stop Charge",
    CMD_ALL_START_DISCHARGE: "All Start Discharge", CMD_ALL_STOP_DISCHARGE: "All Stop Discharge",
    CMD_ALL_SET_CYCLE_NUM: "All Set Cycle Num", CMD_ALL_SET_BAT_TYPE: "All Set Battery Type",
    CMD_ALL_EMERGENCY_STOP: "Emergency Stop",
    CMD_ALL_BUZZER_ON: "All Buzzer On", CMD_ALL_BUZZER_OFF: "All Buzzer Off"
}

STAGE_IDLE = 0
STAGE_CHARGE1 = 1
STAGE_DISCHARGE = 2
STAGE_CHARGE2 = 3
STAGE_COMPLETED = 4
DATA_STALE_RETRY_SEC = 10
DATA_STALE_ABORT_SEC = 60
NO_PROGRESS_ABORT_SEC = 30 * 60
STAGE_HARD_ABORT_SEC = 12 * 60 * 60
TEST_START_FRESH_SEC = 5

CURRENT_THRESHOLD = 50  # mA


# ---------- 双语支持 ----------
class LanguageManager:
    _instance = None
    def __init__(self):
        self.lang = 'zh'  # 'zh' or 'en'
        self.strings = {
            # 通用
            'ok': {'zh': '确定', 'en': 'OK'},
            'cancel': {'zh': '取消', 'en': 'Cancel'},
            'yes': {'zh': '是', 'en': 'Yes'},
            'no': {'zh': '否', 'en': 'No'},
            'warning': {'zh': '警告', 'en': 'Warning'},
            'error': {'zh': '错误', 'en': 'Error'},
            'information': {'zh': '信息', 'en': 'Information'},
            'confirm': {'zh': '确认', 'en': 'Confirm'},
            'clear_log': {'zh': '清空日志', 'en': 'Clear Log'},
            # UI 标题
            'company_name_cn': {'zh': '深圳市家佳明光电科技有限公司', 'en': 'ALITE Co., Ltd.'},
            'company_name_en': {'zh': 'ALITE Co., Ltd.', 'en': 'ALITE Co., Ltd.'},
            'window_title': {'zh': '电池测试仪监控', 'en': 'Battery Tester Monitor'},
            'control_panel': {'zh': '控制面板', 'en': 'Control Panel'},
            'serial_settings': {'zh': '串口设置', 'en': 'Serial Settings'},
            'port': {'zh': '端口:', 'en': 'Port:'},
            'baudrate': {'zh': '波特率:', 'en': 'Baudrate:'},
            'refresh': {'zh': '刷新', 'en': 'Refresh'},
            'connect': {'zh': '连接', 'en': 'Connect'},
            'disconnect': {'zh': '断开', 'en': 'Disconnect'},
            'select_module': {'zh': '选择模块:', 'en': 'Select Module:'},
            'start_charge': {'zh': '开始充电', 'en': 'Start Charge'},
            'stop_charge': {'zh': '停止充电', 'en': 'Stop Charge'},
            'start_discharge': {'zh': '开始放电', 'en': 'Start Discharge'},
            'stop_discharge': {'zh': '停止放电', 'en': 'Stop Discharge'},
            'all_modules_control': {'zh': '全部模块控制', 'en': 'All Modules Control'},
            'all_start_charge': {'zh': '全部开始充电', 'en': 'All Start Charge'},
            'all_stop_charge': {'zh': '全部停止充电', 'en': 'All Stop Charge'},
            'all_start_discharge': {'zh': '全部开始放电', 'en': 'All Start Discharge'},
            'all_stop_discharge': {'zh': '全部停止放电', 'en': 'All Stop Discharge'},
            'emergency_stop': {'zh': '紧急停止 (全部停止)', 'en': 'EMERGENCY STOP (All Stop)'},
            'stop_buzzer': {'zh': '停止蜂鸣器', 'en': 'Stop Buzzer'},
            'set_cycle_num': {'zh': '设置循环次数:', 'en': 'Set Cycle Number:'},
            'set_for_selected': {'zh': '为选中模块设置', 'en': 'Set for Selected'},
            'set_for_all': {'zh': '为所有模块设置', 'en': 'Set for ALL'},
            'battery_type': {'zh': '电池类型 (所有在线模块):', 'en': 'Battery Type (All Online):'},
            'set_for_all_online': {'zh': '为所有在线模块设置', 'en': 'Set for All Online'},
            'export_excel': {'zh': '导出历史记录到Excel', 'en': 'Export History to Excel'},
'single_module': {'zh': '单模块', 'en': 'Single Module'},
            'all_modules': {'zh': '全部模块', 'en': 'All Modules'},
            'settings': {'zh': '设置', 'en': 'Settings'},
            'clear_records': {'zh': '清除记录', 'en': 'Clear Records'},
            'clear_records_confirm': {'zh': '确认清除所选模块的循环记录？', 'en': 'Clear cycle records for selected module?'},
            'quality_pass': {'zh': '合格', 'en': 'PASS'},
            'quality_fail': {'zh': '不合格', 'en': 'FAIL'},
            'no_data': {'zh': '无数据', 'en': 'No Data'},
            'no_battery': {'zh': '无电池', 'en': 'No Battery'},
            'no_battery_short': {'zh': '无电池', 'en': 'NoBat'},
            'executor_offline': {'zh': '中转在线/执行离线', 'en': 'Bridge OK / Executor Offline'},
            'executor_offline_short': {'zh': '执行离线', 'en': 'Exec Offline'},
            'collapse_panel': {'zh': '收起控制面板', 'en': 'Collapse Panel'},
            'expand_panel': {'zh': '展开控制面板', 'en': 'Expand Panel'},
            'test_status_idle': {'zh': '状态: 空闲', 'en': 'Status: Idle'},
            'global_test_status_idle': {'zh': '状态: 空闲', 'en': 'Status: Idle'},
            'bat_3s_li': {'zh': '5节三元锂', 'en': 'Li-Ion 5S'},
            'bat_4s_li': {'zh': '4节三元锂', 'en': 'Li-Ion 4S'},
            'bat_4s_lfp': {'zh': '4节铁锂', 'en': 'LiFePO4 4S'},
            'bat_reserved': {'zh': '保留', 'en': 'Reserved'},
                        'single_module_test': {'zh': '单模块测试', 'en': 'Single Module Test'},
            'lower_limit': {'zh': '下限 (mAh):', 'en': 'Lower Limit (mAh):'},
            'upper_limit': {'zh': '上限 (mAh):', 'en': 'Upper Limit (mAh):'},
            'start_test_on_selected': {'zh': '开始测试选中模块', 'en': 'Start Test on Selected'},
            'stop_test': {'zh': '停止测试', 'en': 'Stop Test'},
            'test_status': {'zh': '测试状态: 空闲', 'en': 'Test Status: Idle'},
            'global_test': {'zh': '全局测试 (所有在线模块)', 'en': 'Global Test (All Online)'},
            'start_loop_test': {'zh': '开始循环测试', 'en': 'Start Loop Test'},
            'stop_loop_test': {'zh': '停止循环测试', 'en': 'Stop Loop Test'},
            'start_global_test': {'zh': '开始全局测试', 'en': 'Start Global Test'},
            'global_test_status_idle': {'zh': '全局测试: 空闲', 'en': 'Global Test: Idle'},
            'current_data': {'zh': '当前数据:', 'en': 'Current Data:'},
            'cycle_history': {'zh': '循环历史记录 (每循环充电/放电容量mAh):', 'en': 'Cycle History (Charge/Discharge mAh per cycle):'},
            'show_comm_log': {'zh': '显示通信日志', 'en': 'Show Communication Log'},
            # 表格列标题
            'col_addr': {'zh': '地址', 'en': 'Addr'},
            'col_voltage': {'zh': '电压(mV)', 'en': 'Voltage(mV)'},
            'col_chg_cur': {'zh': '充电电流(mA)', 'en': 'Chg Cur(mA)'},
            'col_dchg_cur': {'zh': '放电电流(mA)', 'en': 'Dchg Cur(mA)'},
            'col_chg_mah': {'zh': '充电mAh', 'en': 'Chg mAh'},
            'col_dchg_mah': {'zh': '放电mAh', 'en': 'Dchg mAh'},
            'col_status': {'zh': '状态', 'en': 'Status'},
            'col_cycle': {'zh': '循环', 'en': 'Cycle'},
            'col_ck_fail': {'zh': '校验失败', 'en': 'CRC Err'},
            'col_test_status': {'zh': '测试状态', 'en': 'Test Status'},
            'col_qc_result': {'zh': 'QC结果', 'en': 'QC Result'},
            'col_test_caps': {'zh': '测试容量', 'en': 'Test Caps'},
            'col_alarm_type': {'zh': '报警类型', 'en': 'Alarm Type'},
            # 对话框消息
            'no_ports_found': {'zh': '未找到端口', 'en': 'No ports found'},
            'not_connected': {'zh': '未连接', 'en': 'Not connected'},
            'no_online_modules': {'zh': '未检测到在线模块', 'en': 'No online modules detected'},
            'confirm_emergency_stop': {'zh': '发送停止命令给所有模块?', 'en': 'Send STOP to ALL modules?'},
            'confirm_set_bat_type': {'zh': '为所有在线模块设置电池类型为 {}?', 'en': 'Set battery type to {} for all online modules?'},
            'missing_library': {'zh': '缺少库', 'en': 'Missing Library'},
            'install_openpyxl': {'zh': '请安装 openpyxl', 'en': 'Please install openpyxl'},
            'export_complete': {'zh': '导出完成', 'en': 'Export Complete'},
            'export_path': {'zh': '历史记录已导出至:\n{}', 'en': 'History exported to:\n{}'},
            'test_running': {'zh': '测试已在运行中', 'en': 'A test is already running'},
            'confirm_start_global_test': {'zh': '在 {} 个模块上开始独立测试?\n限制: {} - {} mAh', 'en': 'Start independent test on {} module(s)?\nLimits: {} - {} mAh'},
            'confirm_start_test': {'zh': '在模块 {} 上开始测试?', 'en': 'Start test on module {}?'},
            'confirm_stop_test': {'zh': '停止当前测试以及所有充放电?', 'en': 'Stop current test and all charging/discharging?'},
            'test_timeout': {'zh': '测试阶段超时！强制停止。', 'en': 'Test stage timeout! Forcing stop.'},
            'test_stopped_user': {'zh': '测试已停止', 'en': 'Test stopped by user'},
            'confirm_stop_loop': {'zh': '停止当前循环测试?', 'en': 'Stop current loop test?'},
            'loop_completed': {'zh': '所有循环测试完成。', 'en': 'All loop cycles completed.'},
            'loop_completed_msg': {'zh': '循环测试已完成 {} 次循环。\n容量已保存在循环历史中。', 'en': 'Loop test completed after {} cycles.\nCapacities saved in cycle history.'},
            'loop_stopped_user': {'zh': '用户已停止循环测试。', 'en': 'Loop test stopped by user.'},
            'module_no_data': {'zh': '模块 {} 无数据', 'en': 'Module {} has no data'},
            'cycle_num_warning': {'zh': '循环次数至少为1', 'en': 'Cycle number must be at least 1'},
            'confirm_start_loop': {'zh': '在 {} 个模块上开始循环测试?\n总循环次数: {} (每个循环 = 充电 → 放电)', 'en': 'Start loop test on {} module(s)?\nTotal cycles: {} (each cycle = charge → discharge)'},
            'global_test_complete_msg': {'zh': '所有模块测试完成。QC结果显示在“QC结果”列中。', 'en': 'All modules tested. Results shown in QC Result column.'},
            # 状态栏和日志消息
            'connected': {'zh': '已连接到 {} 波特率 {}', 'en': 'Connected to {} at {} baud'},
            'serial_error': {'zh': '串口错误: {}', 'en': 'Serial error: {}'},
            'serial_port_closed': {'zh': '串口已关闭', 'en': 'Serial port closed'},
            'cannot_send': {'zh': '无法发送命令: 串口未打开', 'en': 'Cannot send command: serial port not open'},
            'write_error': {'zh': '写入错误: {}', 'en': 'Write error: {}'},
            'discard_garbage': {'zh': '丢弃无效数据 (无 AA): {}...', 'en': 'Discarding garbage (no AA): {}...'},
            'skip_garbage': {'zh': '跳过 AA 前的垃圾数据: {}...', 'en': 'Skipping garbage before AA: {}...'},
            'checksum_error': {'zh': '校验和错误', 'en': 'Checksum error'},
            'batch_data_len_error': {'zh': '批量数据长度错误', 'en': 'Batch data length mismatch'},
            'command_succeeded': {'zh': "命令 '{}' 成功。", 'en': "Command '{}' succeeded."},
            'command_failed': {'zh': "命令 '{}' 失败。", 'en': "Command '{}' FAILED."},
            'emergency_stop_sent': {'zh': '紧急停止已发送', 'en': 'Emergency stop sent'},
            'buzzers_off': {'zh': '所有蜂鸣器已关闭', 'en': 'All buzzers turned off'},
            'set_cycle_num_module': {'zh': '为 {} 设置循环次数为 {}', 'en': 'Set cycle number for {} to {}'},
            'set_cycle_num_all': {'zh': '为所有模块设置循环次数为 {}', 'en': 'Set cycle number for ALL modules to {}'},
            'set_bat_type_all': {'zh': '为所有模块设置电池类型', 'en': 'Set battery type for all modules'},
            'default_settings_sent': {'zh': '默认设置已发送给所有在线模块', 'en': 'Default settings sent to all online modules'},
            'alarm_overcharge': {'zh': '模块 {} 过充报警！', 'en': 'Module {} OVERCHARGE ALARM!'},
            'alarm_overdischarge': {'zh': '模块 {} 过放报警！', 'en': 'Module {} OVERDISCHARGE ALARM!'},
            'charge_started': {'zh': '模块 {} 开始充电', 'en': 'Module {} charging started'},
            'charge_stopped': {'zh': '模块 {} 停止充电', 'en': 'Module {} charging stopped'},
            'discharge_started': {'zh': '模块 {} 开始放电', 'en': 'Module {} discharging started'},
            'discharge_stopped': {'zh': '模块 {} 停止放电', 'en': 'Module {} discharging stopped'},
            'charge1_completed': {'zh': '模块 {} 第一次充电完成', 'en': 'Module {} Charge1 completed'},
            'discharge_completed': {'zh': '模块 {} 放电完成', 'en': 'Module {} Discharge completed'},
            'discharge_done': {'zh': '模块 {} 放电完成: {}mAh, 开始第二次充电', 'en': 'Module {} discharge done: {}mAh, start charge2'},
            'charge2_completed': {'zh': '模块 {} 第二次充电完成', 'en': 'Module {} Charge2 completed'},
            'charge2_empty_retry': {'zh': '模块 {} 第二次充电容量为0, 重试', 'en': 'Module {} Charge2 empty, retry'},
            'test_finished_result': {'zh': '模块 {} 测试结束, 结果={}', 'en': 'Module {} test finished, result={}'},
            'abort_test_reason': {'zh': '模块 {} 测试中止: {}', 'en': 'Module {} test aborted: {}'},
            'global_test_started': {'zh': '全局独立测试已开始, 模块数: {}', 'en': 'Global independent test started on {} modules'},
            'single_test_started': {'zh': '单模块测试已开始于 {}', 'en': 'Single module test started on {}'},
            'loop_test_started': {'zh': '循环测试已开始, 每个模块目标循环数: {}', 'en': 'Loop test started, target cycles: {} for each module'},
            'module_completed_cycles': {'zh': '模块 {} 已完成 {} 次循环', 'en': 'Module {} completed {} cycles'},
            'all_modules_completed': {'zh': '所有模块已完成当前循环', 'en': 'All modules completed current cycle'},
            'all_modules_completed_loop': {'zh': '所有模块已完成循环测试', 'en': 'All modules completed loop test'},
            'stop_all_modules': {'zh': '已停止所有模块并关闭蜂鸣器', 'en': 'All modules stopped and buzzers off'},
            'test_running_global': {'zh': '全局测试运行中...', 'en': 'Global test running...'},
            'global_test_independent': {'zh': '全局测试: 独立运行', 'en': 'Global Test: Running independently'},
            'single_test_first_charge': {'zh': '测试状态: 首次充电中 {}', 'en': 'Test Status: First Charging on {}'},
            'loop_test_running': {'zh': '循环测试运行中...', 'en': 'Loop test running...'},
            'loop_test_status': {'zh': '循环测试: {}/{} 循环 (进行中)', 'en': 'Loop Test: {}/{} cycles (ongoing)'},
'charge_started_short': {'zh': '充电中 ', 'en': 'Charging '},
            'discharge_started_short': {'zh': '放电中 ', 'en': 'Discharging '},
            'overcharge_alarm_short': {'zh': '过充报警 ', 'en': 'OV Alarm '},
            'overdischarge_alarm_short': {'zh': '过放报警 ', 'en': 'UV Alarm '},
            'charge_complete_short': {'zh': '充电完成 ', 'en': 'Chg Done '},
            'discharge_complete_short': {'zh': '放电完成 ', 'en': 'Dchg Done '},
            'idle_status': {'zh': '空闲', 'en': 'Idle'},
            'charging_status': {'zh': '充电 {0}/{1}', 'en': 'Charging {0}/{1}'},
            'discharging_status': {'zh': '放电 {0}/{1}', 'en': 'Discharging {0}/{1}'},
            'completed_status': {'zh': '完成 {0}/{1}', 'en': 'Done {0}/{1}'},
            'idle_cycle_status': {'zh': '空闲 {0}/{1}', 'en': 'Idle {0}/{1}'},
            'cycle_not_set': {'zh': '未设置循环', 'en': 'Cycle Not Set'},
            'loop_status': {'zh': '循环 {0}/{1}', 'en': 'Loop {0}/{1}'},
            'charge_started_short': {'zh': '充电中 ', 'en': 'Charging '},
            'discharge_started_short': {'zh': '放电中 ', 'en': 'Discharging '},
            'overcharge_alarm_short': {'zh': '过充报警 ', 'en': 'OV Alarm '},
            'overdischarge_alarm_short': {'zh': '过放报警 ', 'en': 'UV Alarm '},
            'charge_complete_short': {'zh': '充电完成 ', 'en': 'Chg Done '},
            'discharge_complete_short': {'zh': '放电完成 ', 'en': 'Dchg Done '},
            'idle_status': {'zh': '空闲', 'en': 'Idle'},
            'charging_status': {'zh': '充电 {0}/{1}', 'en': 'Charging {0}/{1}'},
            'discharging_status': {'zh': '放电 {0}/{1}', 'en': 'Discharging {0}/{1}'},
            'completed_status': {'zh': '完成 {0}/{1}', 'en': 'Done {0}/{1}'},
            'idle_cycle_status': {'zh': '空闲 {0}/{1}', 'en': 'Idle {0}/{1}'},
            'cycle_not_set': {'zh': '未设置循环', 'en': 'Cycle Not Set'},
            'loop_status': {'zh': '循环 {0}/{1}', 'en': 'Loop {0}/{1}'},
        }

    @classmethod
    def instance(cls):
        if cls._instance is None:
            cls._instance = LanguageManager()
        return cls._instance

    def tr(self, key, *args, **kwargs):
        text = self.strings.get(key, {}).get(self.lang, key)
        if args:
            try:
                text = text.format(*args)
            except:
                pass
        if kwargs:
            try:
                text = text.format(**kwargs)
            except:
                pass
        return text

    def set_language(self, lang):
        self.lang = lang


def tr(key, *args, **kwargs):
    return LanguageManager.instance().tr(key, *args, **kwargs)


# ---------- 模块状态机 ----------
class ModuleTestState:
    def __init__(self, idx, lower_limit, upper_limit):
        self.idx = idx
        self.lower_limit = lower_limit
        self.upper_limit = upper_limit
        self.stage = STAGE_IDLE
        self.charge1_start = 0
        self.charge1_max = 0
        self.discharge_start = 0
        self.discharge_max = 0
        self.charge2_start = 0
        self.charge2_max = 0
        self.discharge_cap = 0
        self.charge2_cap = 0
        self.result = None
        self.completed = False
        self.stage_start_time = time.time()
        self.last_progress_time = self.stage_start_time
        self.last_retry_time = 0
        self._last_chg = 0
        self._last_dchg = 0
        self._chg_stable = 0
        self._dchg_stable = 0

    def start(self, initial_charge_mah, initial_discharge_mah):
        self.stage = STAGE_CHARGE1
        self.charge1_start = 0  # STC8H resets on START_CHARGE
        self.charge1_max = 0
        self.discharge_start = 0  # STC8H resets on START_DISCHARGE
        self.discharge_max = 0
        self.charge2_start = 0
        self.charge2_max = 0
        self.discharge_cap = 0
        self.charge2_cap = 0
        self.result = None
        self.completed = False
        self.stage_start_time = time.time()
        self.last_progress_time = self.stage_start_time
        self.last_retry_time = 0
        self._last_chg = 0
        self._last_dchg = 0
        self._chg_stable = 0
        self._dchg_stable = 0


class LoopModuleState:
    def __init__(self, idx, total_cycles):
        self.idx = idx
        self.total_cycles = total_cycles
        self.current_cycle = 0
        self.stage = STAGE_IDLE
        self.completed = False
        self.discharge_start = 0
        self.charge_start = 0
        self.charge_max = 0
        self.discharge_max = 0
        self.stage_start_time = time.time()
        self.last_progress_time = self.stage_start_time
        self.last_retry_time = 0
        self._last_chg = 0
        self._last_dchg = 0
        self._chg_stable = 0
        self._dchg_stable = 0


    def start(self, init_charge_mah, init_discharge_mah):
        self.stage = STAGE_CHARGE1
        self.charge_start = 0  # STC8H resets on START_CHARGE
        self.discharge_start = 0  # STC8H resets on START_DISCHARGE
        self.charge_max = 0
        self.discharge_max = 0
        self.current_cycle = 0
        self.completed = False
        self.charge_end = 0
        self.stage_start_time = time.time()
        self.last_progress_time = self.stage_start_time
        self.last_retry_time = 0
        self._last_chg = 0
        self._last_dchg = 0
        self._chg_stable = 0
        self._dchg_stable = 0



# ---------- 串口线程 ----------
class SerialThread(QThread):
    data_received = pyqtSignal(list)
    command_response = pyqtSignal(int, int, bytes)
    error_occurred = pyqtSignal(str)
    log_message = pyqtSignal(str, str)  # msg, category: comm|error|info
    ck_fail_update = pyqtSignal(dict)

    def __init__(self, port, baudrate=115200):
        super().__init__()
        self.port = port
        self.baudrate = baudrate
        self.ser = None
        self.running = False
        self.lock = threading.Lock()
        self.request_data = True
        self.batch_interval = 1.0  # match ESP32 1s cache interval
        self.last_batch_time = 0
        self.ck_fail_counts = {}
        self.last_cmd_event = threading.Event()
        self.last_cmd_status = 0

    def run(self):
        try:
            self.ser = serial.Serial(self.port, self.baudrate, timeout=0.2)
            self.ser.set_buffer_size(rx_size=8192, tx_size=4096)
            self.log_message.emit(tr('connected', self.port, self.baudrate), "comm")
            self.running = True
            self.ser.reset_input_buffer()
            self.ser.reset_output_buffer()

            while self.running:
                if self.ser.in_waiting:
                    self.read_frame()
                else:
                    time.sleep(0.01)
        except serial.SerialException as e:
            self.error_occurred.emit(tr('serial_error', str(e)))
            self.log_message.emit(tr('serial_error', str(e)), "error")
        except Exception as e:
            self.error_occurred.emit(str(e))
            self.log_message.emit(f"Serial thread error: {e}", "error")
        finally:
            if self.ser and self.ser.is_open:
                self.ser.close()
                self.log_message.emit(tr('serial_port_closed'), "comm")

    def send_command(self, cmd, addr, data=None):
        if self.ser is None or not self.ser.is_open:
            self.log_message.emit(tr('cannot_send'))
            return
        if data is None:
            data = []
        length = len(data)
        frame = bytearray([FRAME_HEADER, cmd, addr])
        frame.append((length >> 8) & 0xFF)
        frame.append(length & 0xFF)
        frame.extend(data)
        checksum = cmd ^ addr ^ (length >> 8) ^ (length & 0xFF)
        for b in data:
            checksum ^= b
        frame.append(checksum)

        with self.lock:
            try:
                self.ser.write(frame)
            except serial.SerialException as e:
                self.log_message.emit(tr('write_error', str(e)), "error")
                self.error_occurred.emit(str(e))
                return

        frame_hex = ' '.join(f'{b:02X}' for b in frame)
        cmd_name = CMD_NAMES.get(cmd, f"Unknown(0x{cmd:02X})")
        addr_str = f"0x{addr:02X}" if addr != 0 else "Broadcast"
        self.log_message.emit(f"[TX] {frame_hex} | Cmd: {cmd_name} | Addr: {addr_str} | Data: {data}", "comm")

    def send_all_command(self, cmd, data=None):
        self.send_command(cmd, 0x00, data)

    def send_batch_read(self):
        self.send_command(CMD_BATCH_READ, 0x00)

    def _scan_text_garbage(self, data):
        """Extract and count CKSUM FAIL messages from serial garbage bytes."""
        try:
            text = data.decode('ascii', errors='ignore')
            for m in re.finditer(r'CKSUM FAIL addr=0x([0-9A-Fa-f]{2})', text):
                addr = int(m.group(1), 16)
                count = self.ck_fail_counts.get(addr, 0) + 1
                self.ck_fail_counts[addr] = count
        except Exception:
            pass

    def _try_decode_module_data(self, raw):
        """Decode raw bytes as module data for human-readable discard logs."""
        parts = []
        for i in range(0, len(raw) - 1, 12):
            chunk = raw[i:i+12]
            if len(chunk) < 12:
                break
            volt = (chunk[0] << 8) | chunk[1]
            chg_mah = (chunk[6] << 8) | chunk[7]
            dchg_mah = (chunk[8] << 8) | chunk[9]
            status = chunk[10]
            cycle = chunk[11]
            if volt > 0 or chg_mah > 0 or dchg_mah > 0:
                parts.append(f"V={volt}mV C={chg_mah}mAh D={dchg_mah}mAh st={status:02X} cy={cycle}")
        return ' | '.join(parts) if parts else ''

    def read_frame(self):
        buffer = bytearray()
        while self.running:
            if self.ser.in_waiting:
                chunk = self.ser.read(self.ser.in_waiting)
                if chunk:
                    buffer.extend(chunk)
            if len(buffer) < 5:
                time.sleep(0.005)
                continue
            idx = buffer.find(b'\xAA')
            if idx == -1:
                if len(buffer) < 400:
                    time.sleep(0.005)
                    continue
                if buffer:
                    self._scan_text_garbage(bytes(buffer))
                    discard = buffer[:min(20, len(buffer))]
                    hex_discard = ' '.join(f'{b:02X}' for b in discard)
                    decoded = self._try_decode_module_data(bytes(buffer))
                    if decoded:
                        hex_discard += ' | ' + decoded[:300]
                    self.log_message.emit(tr('discard_garbage', hex_discard), "error")
                    buffer = buffer[400:] if len(buffer) > 400 else bytearray()
                continue
            if idx > 0:
                self._scan_text_garbage(bytes(buffer[:idx]))
                discard = buffer[:idx]
                hex_discard = ' '.join(f'{b:02X}' for b in discard[:20])
                decoded = self._try_decode_module_data(bytes(discard))
                if decoded:
                    hex_discard += ' | ' + decoded[:300]
                self.log_message.emit(tr('skip_garbage', hex_discard), "error")
                buffer = buffer[idx:]
            if len(buffer) < 5:
                continue
            cmd = buffer[1]
            status = buffer[2]
            len_h = buffer[3]
            len_l = buffer[4]
            data_len = (len_h << 8) | len_l
            total_frame_len = 1 + 4 + data_len + 1
            if len(buffer) < total_frame_len:
                time.sleep(0.005)
                continue
            frame = buffer[:total_frame_len]
            buffer = buffer[total_frame_len:]
            calc_checksum = cmd ^ status ^ len_h ^ len_l
            data_part = frame[5:5+data_len]
            for b in data_part:
                calc_checksum ^= b
            if calc_checksum != frame[-1]:
                self.log_message.emit(tr('checksum_error'), "error")
                buffer = buffer[1:] if buffer else bytearray()
                continue
            self.process_frame(frame, cmd, status, data_len, data_part)


    def send_command_sync(self, cmd, addr, data=None, timeout_ms=300):
        """Send command and wait for ESP32 ACK. Returns: 0=OK, 1=FAIL, -1=timeout"""
        import time as _time
        if self.ser is None or not self.ser.is_open:
            return -1
        self.last_cmd_event.clear()

        # Build frame inline to avoid reentrant lock with send_command()
        if data is None:
            data = []
        length = len(data)
        frame = bytearray([FRAME_HEADER, cmd, addr])
        frame.append((length >> 8) & 0xFF)
        frame.append(length & 0xFF)
        frame.extend(data)
        checksum = cmd ^ addr ^ (length >> 8) ^ (length & 0xFF)
        for b in data:
            checksum ^= b
        frame.append(checksum)

        with self.lock:
            try:
                self.ser.write(frame)
            except serial.SerialException as e:
                self.log_message.emit(tr('write_error', str(e)), "error")
                return -1

        # Log TX
        frame_hex = ' '.join(f'{b:02X}' for b in frame)
        cmd_name = CMD_NAMES.get(cmd, f"Unknown(0x{cmd:02X})")
        addr_str = f"0x{addr:02X}" if addr != 0 else "Broadcast"
        self.log_message.emit(f"[TX] {frame_hex} | Cmd: {cmd_name} | Addr: {addr_str} | Data: {data}", "comm")

        if self.last_cmd_event.wait(timeout_ms / 1000.0):
            return self.last_cmd_status
        return -1

    def process_frame(self, frame, cmd, status, data_len, data):
        frame_hex = 'AA ' + ' '.join(f'{b:02X}' for b in frame[1:])
        cmd_name = CMD_NAMES.get(cmd, f"Unknown(0x{cmd:02X})")
        status_str = 'OK' if status == 0 else 'Error'
        self.log_message.emit(f"[RX] {frame_hex} | Cmd: {cmd_name} | Status: {status_str}", "comm")

        if cmd == CMD_BATCH_READ and status == 0:
            expected_len = MODULE_COUNT * 12
            if data_len != expected_len:
                self.log_message.emit(tr('batch_data_len_error'), "error")
                return
            modules = []
            offset = 0
            for i in range(MODULE_COUNT):
                if offset + 12 > len(data):
                    break
                volt = struct.unpack('>H', data[offset:offset+2])[0]
                chg_cur = struct.unpack('>H', data[offset+2:offset+4])[0]
                dchg_cur = struct.unpack('>H', data[offset+4:offset+6])[0]
                chg_mah = struct.unpack('>H', data[offset+6:offset+8])[0]
                dchg_mah = struct.unpack('>H', data[offset+8:offset+10])[0]
                status_byte = data[offset+10]
                cycle = data[offset+11]
                modules.append({
                    'addr': MODULE_ADDRS[i],
                    'voltage': volt,
                    'charge_current': chg_cur,
                    'charge_mah': chg_mah,
                    'discharge_current': dchg_cur,
                    'discharge_mah': dchg_mah,
                    'status': status_byte,
                    'cycle': cycle,
                    'ck_fail_count': self.ck_fail_counts.get(MODULE_ADDRS[i], 0)
                })
                offset += 12
            self.data_received.emit(modules)
        else:
            self.last_cmd_status = status
            self.last_cmd_event.set()
            self.command_response.emit(cmd, status, bytes(data))

    def stop(self):
        self.running = False
        self.wait()


# ---------- 日志窗口 ----------
class LogWindow(QDialog):
    closed = pyqtSignal()
    CAT_COLORS = {"comm": QColor(100,100,100), "error": QColor(200,30,30), "info": QColor(0,0,0)}

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(tr('show_comm_log'))
        self.resize(750, 450)

        layout = QVBoxLayout(self)

        film = QHBoxLayout()
        self.cb_comm = QCheckBox("[TX/RX]")
        self.cb_comm.setChecked(True)
        self.cb_comm.toggled.connect(self._on_filter_changed)
        film.addWidget(self.cb_comm)
        self.cb_error = QCheckBox("Error")
        self.cb_error.setChecked(True)
        self.cb_error.toggled.connect(self._on_filter_changed)
        film.addWidget(self.cb_error)
        self.cb_info = QCheckBox("Info")
        self.cb_info.setChecked(True)
        self.cb_info.toggled.connect(self._on_filter_changed)
        film.addWidget(self.cb_info)
        film.addStretch()
        self.count_label = QLabel()
        film.addWidget(self.count_label)
        layout.addLayout(film)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setLineWrapMode(QTextEdit.NoWrap)
        layout.addWidget(self.log_text)

        bbar = QHBoxLayout()
        self.btn_clear = QPushButton(tr('clear_log'))
        self.btn_clear.clicked.connect(self._clear_log)
        bbar.addWidget(self.btn_clear)
        bbar.addStretch()
        layout.addLayout(bbar)

        self.max_lines = 1000
        self._all_lines = []
        self._update_count()

    def _on_filter_changed(self):
        self._update_count()
        self._rebuild()

    def _update_count(self):
        v = sum(1 for _, _, c in self._all_lines if self._vis(c))
        self.count_label.setText("%d/%d" % (v, len(self._all_lines)))

    def _vis(self, cat):
        if cat == "comm": return self.cb_comm.isChecked()
        if cat == "error": return self.cb_error.isChecked()
        return self.cb_info.isChecked()

    def _rebuild(self):
        self.log_text.clear()
        for ts, msg, cat in self._all_lines:
            if self._vis(cat):
                self.log_text.setTextColor(self.CAT_COLORS.get(cat, QColor(0,0,0)))
                self.log_text.append("[%s] %s" % (ts, msg))
        self.log_text.ensureCursorVisible()

    def append_log(self, msg, category="info"):
        ts = QTime.currentTime().toString("hh:mm:ss.zzz")
        self._all_lines.append((ts, msg, category))
        while len(self._all_lines) > self.max_lines:
            self._all_lines.pop(0)
        if self._vis(category):
            self.log_text.setTextColor(self.CAT_COLORS.get(category, QColor(0,0,0)))
            self.log_text.append("[%s] %s" % (ts, msg))
            doc = self.log_text.document()
            while doc.blockCount() > self.max_lines:
                c = QTextCursor(doc.findBlockByLineNumber(0))
                c.select(QTextCursor.BlockUnderCursor)
                c.removeSelectedText()
                c.deleteChar()
        self._update_count()
        self.log_text.ensureCursorVisible()

    def _clear_log(self):
        self._all_lines.clear()
        self.log_text.clear()
        self._update_count()

    def closeEvent(self, event):
        self.closed.emit()
        super().closeEvent(event)


# ---------- 主窗口 ----------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(tr('window_title'))
        self.setMinimumSize(1250, 850)

        # 数据变量
        self.modules_data = [None] * MODULE_COUNT
        self.set_cycle_nums = [0] * MODULE_COUNT
        self.cycle_records = [[] for _ in range(MODULE_COUNT)]
        self.last_cycle = [0] * MODULE_COUNT
        self.last_status = [0] * MODULE_COUNT
        self.last_charge_mah = [0] * MODULE_COUNT
        self.last_discharge_mah = [0] * MODULE_COUNT
        self.module_last_seen = [0] * MODULE_COUNT
        self.module_miss_count = [0] * MODULE_COUNT

        self.charge_max = [0] * MODULE_COUNT
        self.discharge_max = [0] * MODULE_COUNT
        self.charge_start_mah = [0] * MODULE_COUNT
        self.discharge_start_mah = [0] * MODULE_COUNT

        self.online_modules = set()
        self.batch_read_received = False

        self.serial_thread = None
        self.log_window = None

        self.test_running = False
        self.global_test_mode = False
        self.module_test_states = {}
        self.global_test_results = [None] * MODULE_COUNT
        self.single_test_state = None

        self.loop_test_active = False
        self.loop_current_cycle = 0
        self.loop_total_cycles = 0

        self.test_stage_timer = QTimer()
        self.test_stage_timer.setSingleShot(True)
        self.test_stage_timer.timeout.connect(self.on_test_timeout)

        self.buzzer_timer = QTimer()
        self.buzzer_timer.timeout.connect(self.check_ng_and_buzz)
        self.buzzer_timer_interval = 5000
        self.ng_buzzer_set = set()
        self.last_alarm_time = 0

        self.current_check_timer = QTimer()
        self.current_check_timer.timeout.connect(self.check_remaining_current)
        self.current_check_timer.setInterval(1000)


        self.batch_read_timer = QTimer()
        self.batch_read_timer.timeout.connect(self.trigger_batch_read)

        self.setup_ui()
        self.refresh_serial_ports()
        self.apply_stylesheet()
        self.retranslate_ui()

        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)

        screen = QApplication.primaryScreen()
        if screen:
            size = screen.availableSize()
            self.resize(int(size.width() * 0.8), int(size.height() * 0.85))

        # 语言切换按钮
        self.lang_btn = QPushButton("English")
        self.lang_btn.clicked.connect(self.toggle_language)
        self.statusBar.addPermanentWidget(self.lang_btn)

    def toggle_left_panel(self):
        if self.left_panel_visible:
            self.left_panel.hide()
            self.btn_collapse.setText(">")
            self.btn_collapse.setToolTip(self.tr('expand_panel'))
            self.left_panel_visible = False
        else:
            self.left_panel.show()
            self.btn_collapse.setText("<")
            self.btn_collapse.setToolTip(self.tr('collapse_panel'))
            self.left_panel_visible = True
        self.update_module_detail()
    
    def toggle_language(self):
        lang_mgr = LanguageManager.instance()
        if lang_mgr.lang == 'zh':
            lang_mgr.set_language('en')
            self.lang_btn.setText("中文")
        else:
            lang_mgr.set_language('zh')
            self.lang_btn.setText("English")
        self.retranslate_ui()

    def retranslate_ui(self):
        self.setWindowTitle(tr('window_title'))
        self.group_control.setTitle(tr('control_panel'))
        self.lbl_single.setText(tr('single_module') + ":")
        self.lbl_all.setText(tr('all_modules') + ":")
        self.label_port.setText("  " + tr('port') + " ")
        self.label_baud.setText(" " + tr('baudrate') + " ")
        self.btn_refresh.setText(tr('refresh'))
        if self.serial_thread:
            self.btn_connect.setText(tr('disconnect'))
        else:
            self.btn_connect.setText(tr('connect'))
        self.label_select_module.setText(" " + tr('select_module') + " ")
        self.btn_start_charge.setText(tr('start_charge'))
        self.btn_stop_charge.setText(tr('stop_charge'))
        self.btn_start_discharge.setText(tr('start_discharge'))
        self.btn_stop_discharge.setText(tr('stop_discharge'))
        self.btn_all_start_charge.setText(tr('all_start_charge'))
        self.btn_all_stop_charge.setText(tr('all_stop_charge'))
        self.btn_all_start_discharge.setText(tr('all_start_discharge'))
        self.btn_all_stop_discharge.setText(tr('all_stop_discharge'))
        self.btn_emergency_stop.setText(tr('emergency_stop'))
        self.btn_stop_buzzer.setText(tr('stop_buzzer'))
        self.label_cycle.setText(tr('set_cycle_num') + ":")
        self.btn_set_cycle.setText(tr('set_for_selected'))
        self.btn_set_cycle_all.setText(tr('set_for_all'))
        self.label_bat_type.setText(tr('battery_type') + ":")
        self.btn_set_bat_type_all.setText(tr('set_for_all_online'))
        self.group_settings.setTitle(tr('settings'))
        self.btn_export_excel.setText(tr('export_excel'))
        self.btn_clear_records.setText(tr('clear_records'))
        self.group_test.setTitle(tr('single_module_test'))
        self.label_lower.setText(tr('lower_limit') + ":")
        self.label_upper.setText(tr('upper_limit') + ":")
        self.btn_start_test.setText(tr('start_test_on_selected'))
        self.btn_stop_test.setText(tr('stop_test'))
        if not self.test_running:
            self.test_status_label.setText(tr('test_status'))
        else:
            if self.loop_test_active:
                self.test_status_label.setText(tr('loop_test_running'))
            elif self.global_test_mode:
                self.test_status_label.setText(tr('test_running_global'))
        self.group_global_test.setTitle(tr('global_test'))
        self.btn_loop_test.setText(tr('start_loop_test'))
        self.btn_stop_loop.setText(tr('stop_loop_test'))
        self.btn_global_test.setText(tr('start_global_test'))
        if not self.loop_test_active and not self.global_test_mode:
            self.global_test_status_label.setText(tr('global_test_status_idle'))
        elif self.loop_test_active:
            self.global_test_status_label.setText(tr('loop_test_status', self.loop_current_cycle, self.loop_total_cycles))
        else:
            self.global_test_status_label.setText(tr('global_test_independent'))
        self.label_current_data.setText(tr('current_data'))
        self.label_history.setText(tr('cycle_history'))
        self.btn_show_log.setText(tr('show_comm_log'))

        # 表格列标题
        headers = [tr('col_addr'), tr('col_voltage'), tr('col_chg_cur'), tr('col_dchg_cur'),
                   tr('col_chg_mah'), tr('col_dchg_mah'), tr('col_status'), tr('col_cycle'),
                   tr('col_test_status'), tr('col_qc_result'), tr('col_test_caps'), tr('col_alarm_type'),
                   tr('col_ck_fail')]
        self.main_table.setHorizontalHeaderLabels(headers)
        current_port = self.port_combo.currentText()          # 保存当前选中的端口名
        self.refresh_serial_ports()                           # 重新扫描并填充端口列表
        index = self.port_combo.findText(current_port)
        if index >= 0:
            self.port_combo.setCurrentIndex(index)
        self.update_table()
        self.update_history_table()

    def apply_stylesheet(self):
        self.setStyleSheet("""
        * { font-family: "Microsoft YaHei", "Segoe UI", sans-serif; font-size: 12px; }
        QMainWindow { background-color: #f0f0f0; }
        QWidget#header { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #ffffff, stop:1 #e8ecf1); border-bottom: 1px solid #c0c0c0; }
        QLabel#logo { background: #e0e0e0; border: 1px dashed #999; border-radius: 4px; font-size: 10px; color: #888; }
        QLabel#company_name_cn { font-size: 16px; font-weight: bold; color: #003399; }
        QLabel#company_name_en { font-size: 11px; color: #666; }
        QGroupBox {
            background: #ffffff; border: 1px solid #d0d0d0;
            border-radius: 4px; margin-top: 10px; padding: 10px 8px 6px 8px;
            font-weight: bold;
        }
        QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 4px; color: #003399; }
        QPushButton {
            background-color: #0066cc; color: #ffffff; border: 1px solid #0055aa;
            border-radius: 3px; padding: 5px 14px; min-height: 24px; font-weight: bold;
        }
        QPushButton:hover { background-color: #0055aa; }
        QPushButton:pressed { background-color: #004488; }
        QPushButton:disabled { background-color: #cccccc; color: #888888; border-color: #bbbbbb; }
        QPushButton#emergency { background-color: #cc0000; border-color: #aa0000; }
        QPushButton#emergency:hover { background-color: #aa0000; }
        QComboBox, QSpinBox {
            background: #ffffff; border: 1px solid #b0b0b0;
            border-radius: 3px; padding: 4px 6px; min-height: 22px;
        }
        QComboBox:hover, QSpinBox:hover { border-color: #0066cc; }
        QComboBox::drop-down { border: none; width: 18px; }
        QTableWidget {
            background: #ffffff; border: 1px solid #c0c0c0;
            gridline-color: #e0e0e0; alternate-background-color: #f5f5f5;
        }
        QTableWidget::item { padding: 2px 4px; }
        QTableWidget::item:selected { background: #cce5ff; color: #000000; }
        QHeaderView::section {
            background-color: #e8e8e8; color: #333333;
            border: 1px solid #d0d0d0; padding: 4px; font-weight: bold;
        }
        QToolBar {
            background-color: #f0f0f0; border: none;
            border-bottom: 1px solid #d0d0d0; spacing: 4px; padding: 2px 6px;
        }
        QToolBar QComboBox { min-width: 70px; }
        QToolBar QPushButton { min-height: 22px; padding: 3px 10px; }
        QWidget#detailPanel { background: #e8f0fe; border: 1px solid #b8d4f0; }
        QStatusBar { background: #e0e0e0; font-size: 11px; }
        QStatusBar::item { border: none; }
        """)



    def setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ===== HEADER =====
        header = QWidget()
        header.setObjectName("header")
        hl = QHBoxLayout(header)
        hl.setContentsMargins(12, 6, 12, 6)
        hl.setSpacing(12)

        self.logo_label = QLabel()
        self.logo_label.setFixedSize(200, 50)
        self.logo_label.setObjectName("logo")
        self.logo_label.setAlignment(Qt.AlignCenter)
        hl.addWidget(self.logo_label)

        name_col = QVBoxLayout()
        name_col.setSpacing(2)
        self.company_name_cn_label = QLabel(tr('company_name_cn'))
        self.company_name_cn_label.setObjectName("company_name_cn")
        name_col.addWidget(self.company_name_cn_label)
        self.company_name_en_label = QLabel(tr('company_name_en'))
        self.company_name_en_label.setObjectName("company_name_en")
        name_col.addWidget(self.company_name_en_label)
        hl.addLayout(name_col)

        hl.addStretch()
        root.addWidget(header)
        self._load_logo()

        # ===== TOOLBAR =====
        tb = QToolBar()
        tb.setMovable(False)

        self.label_port = QLabel("  " + self.tr('port') + " ")
        tb.addWidget(self.label_port)
        self.port_combo = QComboBox()
        self.port_combo.setFixedWidth(80)
        tb.addWidget(self.port_combo)

        self.label_baud = QLabel(" " + self.tr('baudrate') + " ")
        tb.addWidget(self.label_baud)
        self.baud_combo = QComboBox()
        self.baud_combo.addItems(["9600", "19200", "38400", "57600", "115200", "256000"])
        self.baud_combo.setCurrentText("256000")
        self.baud_combo.setFixedWidth(70)
        tb.addWidget(self.baud_combo)

        self.btn_refresh = QPushButton(self.tr('refresh'))
        self.btn_refresh.clicked.connect(self.refresh_serial_ports)
        tb.addWidget(self.btn_refresh)

        self.btn_connect = QPushButton(self.tr('connect'))
        self.btn_connect.clicked.connect(self.on_connect)
        tb.addWidget(self.btn_connect)
        tb.addSeparator()

        self.label_select_module = QLabel(" " + self.tr('select_module') + " ")
        tb.addWidget(self.label_select_module)
        self.module_combo = QComboBox()
        self.module_combo.addItems(["Module %d" % (i+1) for i in range(MODULE_COUNT)])
        self.module_combo.setFixedWidth(100)
        self.module_combo.currentIndexChanged.connect(self.update_module_detail)
        tb.addWidget(self.module_combo)
        tb.addSeparator()

        self.btn_emergency_stop = QPushButton(self.tr('emergency_stop'))
        self.btn_emergency_stop.setObjectName("emergency")
        self.btn_emergency_stop.clicked.connect(self.on_emergency_stop)
        tb.addWidget(self.btn_emergency_stop)

        self.btn_stop_buzzer = QPushButton(self.tr('stop_buzzer'))
        self.btn_stop_buzzer.clicked.connect(self.on_stop_buzzer)
        tb.addWidget(self.btn_stop_buzzer)

        self.btn_collapse = QPushButton("<")
        self.btn_collapse.setToolTip(self.tr('collapse_panel'))
        self.btn_collapse.setFixedWidth(28)
        self.btn_collapse.clicked.connect(self.toggle_left_panel)
        self.left_panel_visible = True
        tb.addWidget(self.btn_collapse)

        root.addWidget(tb)

        # ===== MAIN BODY =====
        body = QHBoxLayout()
        body.setContentsMargins(4, 4, 4, 4)
        body.setSpacing(4)

        # ── LEFT ──
        left_wrap = QWidget()
        left = QVBoxLayout(left_wrap)
        left.setContentsMargins(0, 0, 0, 0)
        left.setSpacing(4)

        # G1: 容量测试
        self.group_test = QGroupBox(self.tr('single_module_test'))
        g1 = QVBoxLayout(self.group_test)
        g1.setSpacing(3)

        lim = QHBoxLayout()
        self.label_lower = QLabel(self.tr('lower_limit') + ":")
        lim.addWidget(self.label_lower)
        self.lower_limit_spin = QSpinBox()
        self.lower_limit_spin.setRange(0, 100000)
        self.lower_limit_spin.setValue(1800)
        lim.addWidget(self.lower_limit_spin)
        self.label_upper = QLabel(self.tr('upper_limit') + ":")
        lim.addWidget(self.label_upper)
        self.upper_limit_spin = QSpinBox()
        self.upper_limit_spin.setRange(0, 100000)
        self.upper_limit_spin.setValue(2200)
        lim.addWidget(self.upper_limit_spin)
        g1.addLayout(lim)

        tbtns = QHBoxLayout()
        self.btn_start_test = QPushButton(self.tr('start_test_on_selected'))
        self.btn_start_test.clicked.connect(self.on_start_test)
        tbtns.addWidget(self.btn_start_test)
        self.btn_stop_test = QPushButton(self.tr('stop_test'))
        self.btn_stop_test.clicked.connect(self.on_stop_test_clicked)
        self.btn_stop_test.setEnabled(False)
        tbtns.addWidget(self.btn_stop_test)
        g1.addLayout(tbtns)

        self.btn_global_test = QPushButton(self.tr('start_global_test'))
        self.btn_global_test.clicked.connect(self.on_start_global_test)
        g1.addWidget(self.btn_global_test)

        self.test_status_label = QLabel(self.tr('test_status_idle'))
        g1.addWidget(self.test_status_label)
        left.addWidget(self.group_test)

        # G2: 循环测试
        self.group_global_test = QGroupBox(self.tr('global_test'))
        g2 = QVBoxLayout(self.group_global_test)
        g2.setSpacing(3)

        cyc = QHBoxLayout()
        self.label_cycle = QLabel(self.tr('set_cycle_num') + ":")
        cyc.addWidget(self.label_cycle)
        self.cycle_spin = QSpinBox()
        self.cycle_spin.setRange(1, 255)
        self.cycle_spin.setValue(10)
        cyc.addWidget(self.cycle_spin)
        self.btn_set_cycle = QPushButton(self.tr('set_for_selected'))
        self.btn_set_cycle.clicked.connect(self.on_set_cycle_num)
        cyc.addWidget(self.btn_set_cycle)
        self.btn_set_cycle_all = QPushButton(self.tr('set_for_all'))
        self.btn_set_cycle_all.clicked.connect(self.on_set_cycle_num_all)
        cyc.addWidget(self.btn_set_cycle_all)
        g2.addLayout(cyc)

        lb = QHBoxLayout()
        self.btn_loop_test = QPushButton(self.tr('start_loop_test'))
        self.btn_loop_test.clicked.connect(self.on_start_loop_test)
        lb.addWidget(self.btn_loop_test)
        self.btn_stop_loop = QPushButton(self.tr('stop_loop_test'))
        self.btn_stop_loop.clicked.connect(self.on_stop_loop_test)
        self.btn_stop_loop.setEnabled(False)
        lb.addWidget(self.btn_stop_loop)
        g2.addLayout(lb)

        self.global_test_status_label = QLabel(self.tr('global_test_status_idle'))
        g2.addWidget(self.global_test_status_label)
        left.addWidget(self.group_global_test)

        # G3: 手动控制
        self.group_control = QGroupBox(self.tr('control_panel'))
        g3 = QGridLayout(self.group_control)
        g3.setSpacing(2)

        self.lbl_single = QLabel(self.tr('single_module') + ":")
        g3.addWidget(self.lbl_single, 0, 0, 1, 2)
        self.btn_start_charge = QPushButton(self.tr('start_charge'))
        self.btn_start_charge.clicked.connect(lambda: self.on_single_command(CMD_START_CHARGE))
        g3.addWidget(self.btn_start_charge, 1, 0)
        self.btn_stop_charge = QPushButton(self.tr('stop_charge'))
        self.btn_stop_charge.clicked.connect(lambda: self.on_single_command(CMD_STOP_CHARGE))
        g3.addWidget(self.btn_stop_charge, 1, 1)
        self.btn_start_discharge = QPushButton(self.tr('start_discharge'))
        self.btn_start_discharge.clicked.connect(lambda: self.on_single_command(CMD_START_DISCHARGE))
        g3.addWidget(self.btn_start_discharge, 2, 0)
        self.btn_stop_discharge = QPushButton(self.tr('stop_discharge'))
        self.btn_stop_discharge.clicked.connect(lambda: self.on_single_command(CMD_STOP_DISCHARGE))
        g3.addWidget(self.btn_stop_discharge, 2, 1)

        self.lbl_all = QLabel(self.tr('all_modules') + ":")
        g3.addWidget(self.lbl_all, 3, 0, 1, 2)
        self.btn_all_start_charge = QPushButton(self.tr('all_start_charge'))
        self.btn_all_start_charge.clicked.connect(lambda: self.on_all_command(CMD_START_CHARGE))
        g3.addWidget(self.btn_all_start_charge, 4, 0)
        self.btn_all_stop_charge = QPushButton(self.tr('all_stop_charge'))
        self.btn_all_stop_charge.clicked.connect(lambda: self.on_all_command(CMD_STOP_CHARGE))
        g3.addWidget(self.btn_all_stop_charge, 4, 1)
        self.btn_all_start_discharge = QPushButton(self.tr('all_start_discharge'))
        self.btn_all_start_discharge.clicked.connect(lambda: self.on_all_command(CMD_START_DISCHARGE))
        g3.addWidget(self.btn_all_start_discharge, 5, 0)
        self.btn_all_stop_discharge = QPushButton(self.tr('all_stop_discharge'))
        self.btn_all_stop_discharge.clicked.connect(lambda: self.on_all_command(CMD_STOP_DISCHARGE))
        g3.addWidget(self.btn_all_stop_discharge, 5, 1)
        left.addWidget(self.group_control)

        # G4: 设置
        self.group_settings = QGroupBox(self.tr('settings'))
        g4 = QVBoxLayout(self.group_settings)
        g4.setSpacing(3)

        bat = QHBoxLayout()
        self.label_bat_type = QLabel(self.tr('battery_type') + ":")
        bat.addWidget(self.label_bat_type)
        self.battery_type_combo = QComboBox()
        self.battery_type_combo.addItem(tr('bat_3s_li'), 0x00)
        self.battery_type_combo.addItem(tr('bat_4s_li'), 0x01)
        self.battery_type_combo.addItem(tr('bat_4s_lfp'), 0x02)
        self.battery_type_combo.addItem(tr('bat_reserved'), 0x03)
        bat.addWidget(self.battery_type_combo)
        self.btn_set_bat_type_all = QPushButton(self.tr('set_for_all_online'))
        self.btn_set_bat_type_all.clicked.connect(self.on_set_battery_type_all)
        bat.addWidget(self.btn_set_bat_type_all)
        g4.addLayout(bat)

        btns = QHBoxLayout()
        self.btn_export_excel = QPushButton(self.tr('export_excel'))
        self.btn_export_excel.clicked.connect(self.on_export_excel)
        btns.addWidget(self.btn_export_excel)
        self.btn_clear_records = QPushButton(self.tr('clear_records'))
        self.btn_clear_records.clicked.connect(self.on_clear_records)
        btns.addWidget(self.btn_clear_records)
        self.btn_show_log = QPushButton(self.tr('show_comm_log'))
        self.btn_show_log.clicked.connect(self.show_log_window)
        btns.addWidget(self.btn_show_log)
        g4.addLayout(btns)
        left.addWidget(self.group_settings)

        left.addStretch()

        self.left_panel = QScrollArea()
        self.left_panel.setWidgetResizable(True)
        self.left_panel.setWidget(left_wrap)
        self.left_panel.setFixedWidth(456)
        self.left_panel.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.left_panel.setFrameShape(QFrame.NoFrame)
        body.addWidget(self.left_panel)

        # ── RIGHT ──
        splitter = QSplitter(Qt.Vertical)

        # Main table
        top_w = QWidget()
        top_l = QVBoxLayout(top_w)
        top_l.setContentsMargins(0, 0, 0, 0)
        top_l.setSpacing(2)
        self.label_current_data = QLabel(self.tr('current_data'))
        f = self.label_current_data.font()
        f.setBold(True)
        self.label_current_data.setFont(f)
        top_l.addWidget(self.label_current_data)
        self.main_table = QTableWidget(MODULE_COUNT, 13)
        self.main_table.setAlternatingRowColors(True)
        self.main_table.verticalHeader().setVisible(True)
        for i in range(MODULE_COUNT):
            self.main_table.setVerticalHeaderItem(i, QTableWidgetItem(str(i+1)))
        top_l.addWidget(self.main_table)
        splitter.addWidget(top_w)

        # Detail bar
        self.detail_widget = QWidget()
        self.detail_widget.setObjectName("detailPanel")
        dl = QHBoxLayout(self.detail_widget)
        dl.setContentsMargins(8, 4, 8, 4)
        dl.setSpacing(12)
        self.lbl_detail_addr = QLabel()
        f2 = self.lbl_detail_addr.font()
        f2.setBold(True)
        self.lbl_detail_addr.setFont(f2)
        dl.addWidget(self.lbl_detail_addr)
        dl.addWidget(QLabel("|"))
        self.lbl_detail_status = QLabel()
        dl.addWidget(self.lbl_detail_status)
        dl.addWidget(QLabel("|"))
        self.lbl_detail_cycle = QLabel()
        dl.addWidget(self.lbl_detail_cycle)
        dl.addWidget(QLabel("|"))
        self.lbl_detail_chg = QLabel()
        dl.addWidget(self.lbl_detail_chg)
        dl.addWidget(QLabel("|"))
        self.lbl_detail_dchg = QLabel()
        dl.addWidget(self.lbl_detail_dchg)
        dl.addWidget(QLabel("|"))
        self.lbl_detail_qc = QLabel()
        dl.addWidget(self.lbl_detail_qc)
        dl.addStretch()
        splitter.addWidget(self.detail_widget)

        # History table
        bot_w = QWidget()
        bot_l = QVBoxLayout(bot_w)
        bot_l.setContentsMargins(0, 0, 0, 0)
        bot_l.setSpacing(2)
        self.label_history = QLabel(self.tr('cycle_history'))
        f3 = self.label_history.font()
        f3.setBold(True)
        self.label_history.setFont(f3)
        bot_l.addWidget(self.label_history)
        self.history_table = QTableWidget(MODULE_COUNT, 40)
        hist_headers = ['PrChg']
        for c in range(1, 40):
            n = (c + 1) // 2
            phase = 'Dchg' if c % 2 == 1 else 'Chg'
            hist_headers.append(f'C{n}{phase}')
        self.history_table.setHorizontalHeaderLabels(hist_headers)
        self.history_table.verticalHeader().setVisible(True)
        for i in range(MODULE_COUNT):
            self.history_table.setVerticalHeaderItem(i, QTableWidgetItem(str(i+1)))
        self.history_table.setAlternatingRowColors(True)
        bot_l.addWidget(self.history_table)
        splitter.addWidget(bot_w)

        splitter.setSizes([500, 36, 250])
        body.addWidget(splitter, 1)
        root.addLayout(body)

        # Timers
        self.table_timer = QTimer()
        self.table_timer.timeout.connect(self.update_table)
        self.table_timer.start(1000)

        self.history_timer = QTimer()
        self._auto_save_path = None
        self._auto_save_file = None
        self._peak_chg_mah = [0] * MODULE_COUNT
        self._peak_dchg_mah = [0] * MODULE_COUNT
        self._display = [None] * MODULE_COUNT  # validated display buffer
        self._nobat_debounce = [0] * MODULE_COUNT  # debounce counter for no-battery clearing
        self.history_timer.timeout.connect(self.update_history_table)
        self.history_timer.start(500)

        self.watchdog_timer = QTimer()
        self.watchdog_timer.timeout.connect(self._state_watchdog)

    # ---------- 辅助功能 ----------
    # ---------- 辅助功能 ----------
    # ---------- 辅助功能 ----------
    # ---------- 辅助功能 ----------
    def trigger_batch_read(self):
        if self.serial_thread and self.serial_thread.ser and self.serial_thread.ser.is_open:
            self.serial_thread.send_batch_read()

    def update_module_detail(self):
        idx = self.module_combo.currentIndex()
        if idx < 0 or idx >= MODULE_COUNT:
            return
        mod = self.modules_data[idx]
        addr = MODULE_ADDRS[idx]
        if mod is None:
            self.lbl_detail_addr.setText("Module %d (Addr %d)" % (idx+1, addr))
            self.lbl_detail_status.setText(self.tr("no_data", "No Data"))
            self.lbl_detail_cycle.setText("")
            self.lbl_detail_chg.setText("")
            self.lbl_detail_dchg.setText("")
            self.lbl_detail_qc.setText("")
            return
        if not mod.get("executor_online", True):
            self.lbl_detail_addr.setText("Module %d (0x%02X)" % (idx+1, addr))
            self.lbl_detail_status.setText(self.tr("executor_offline", "Bridge OK / Executor Offline"))
            self.lbl_detail_cycle.setText("")
            self.lbl_detail_chg.setText("")
            self.lbl_detail_dchg.setText("")
            self.lbl_detail_qc.setText("")
            return
        parts = []
        if mod["status"] & 0x01: parts.append("Chg")
        if mod["status"] & 0x02: parts.append("Dchg")
        if mod["status"] & 0x10: parts.append("ChgDone")
        if mod["status"] & 0x20: parts.append("DchgDone")
        if mod["status"] & 0x04: parts.append("OV!")
        if mod["status"] & 0x08: parts.append("UV!")
        st = " ".join(parts) if parts else "Idle"
        cyc = mod.get("cycle", 0)
        sc = self.set_cycle_nums[idx]
        cs = "Cyc %d/%d" % (cyc, sc) if sc > 0 else "Cyc %d" % cyc
        chg = mod.get("chg1_mah", 0) or mod.get("chg2_mah", 0) or 0
        dchg = mod.get("disc_mah", 0) or 0
        if mod.get("chg2_mah") is not None:
            r = mod["chg2_mah"]
            lo = self.lower_limit_spin.value()
            hi = self.upper_limit_spin.value()
            if lo <= r <= hi:
                self.lbl_detail_qc.setText('<span style="color:#16a34a;font-weight:bold;">%s</span>' % self.tr("quality_pass", "PASS"))
            else:
                self.lbl_detail_qc.setText('<span style="color:#dc2626;font-weight:bold;">%s</span>' % self.tr("quality_fail", "FAIL"))
        else:
            self.lbl_detail_qc.setText("")
        self.lbl_detail_addr.setText("Module %d (0x%02X)" % (idx+1, addr))
        self.lbl_detail_status.setText(st)
        self.lbl_detail_cycle.setText(cs)
        self.lbl_detail_chg.setText("Chg: %d mAh" % chg)
        self.lbl_detail_dchg.setText("Dchg: %d mAh" % dchg)

    def on_clear_records(self):
        reply = QMessageBox.question(
            self, self.tr("clear_records", "Clear Records"),
            self.tr("clear_records_confirm", "Clear cycle records for selected module?"),
            QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            idx = self.module_combo.currentIndex()
            if 0 <= idx < MODULE_COUNT:
                self.cycle_records[idx].clear()
                self.update_history_table()


    def _start_auto_save(self):
        # Close previous file if double-called (safety)
        if self._auto_save_file:
            try:
                self._auto_save_file.close()
            except:
                pass
            self._auto_save_file = None
        from datetime import datetime as _dt
        import os
        savedir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '测试数据')
        try:
            os.makedirs(savedir, exist_ok=True)
            fn = _dt.now().strftime('autosave_%Y%m%d_%H%M%S.csv')
            self._auto_save_path = os.path.join(savedir, fn)
            self._auto_save_file = open(self._auto_save_path, 'w', encoding='utf-8-sig')
            self._auto_save_file.write('时间,模块编号,地址,类型,容量mAh,循环\n')
            self._auto_save_file.flush()
            os.fsync(self._auto_save_file.fileno())
            self.on_log_message('Auto-save started: ' + fn, 'info')
            return True
        except Exception as e:
            self._auto_save_file = None
            self.on_log_message('Auto-save start failed: ' + str(e), 'error')
            QMessageBox.critical(self, tr('error'), 'Auto-save start failed: ' + str(e))
            return False

    def _stop_auto_save(self):
        if self._auto_save_file:
            try:
                self._auto_save_file.flush()
                self._auto_save_file.close()
            except Exception as e:
                self.on_log_message('Auto-save close error: ' + str(e), 'error')
            self._auto_save_file = None
            self.on_log_message('Auto-save stopped: ' + (self._auto_save_path or ''), 'info')

    def _write_save_line(self, module_idx, field, value, voltage=None):
        if (value <= 0 and not str(field).startswith('fail')) or not self._auto_save_file:
            return
        from datetime import datetime as _dt
        import os
        addr = MODULE_ADDRS[module_idx]
        recs = self.cycle_records[module_idx]
        cycle_num = len(recs)
        vstr = (',' + str(voltage)) if voltage is not None else ''
        line = _dt.now().strftime('%Y-%m-%d %H:%M:%S') + ',Module' + str(module_idx+1) + ',' + str(addr) + ',' + field + ',' + str(value) + ',' + str(cycle_num) + vstr + '\n'
        try:
            self._auto_save_file.write(line)
            self._auto_save_file.flush()
            os.fsync(self._auto_save_file.fileno())
        except Exception as e:
            self.on_log_message('Auto-save write error: ' + str(e), 'error')

    def add_cycle_record(self, module_idx, field, value, voltage=None):
        recs = self.cycle_records[module_idx]
        addr = MODULE_ADDRS[module_idx]
        if field == 'charge':
            recs.append({'charge': value, 'discharge': None, 'chg_volt': voltage, 'dchg_volt': None})
            vstr = (' V=%dmV' % voltage) if voltage else ''
            self.on_log_message("RECORD[%#04x] charge = %d mAh%s | records=%d" % (addr, value, vstr, len(recs)), "info")
            self._write_save_line(module_idx, 'charge', value, voltage)
        else:
            if not recs:
                recs.append({'charge': None, 'discharge': value, 'chg_volt': None, 'dchg_volt': voltage})
            else:
                recs[-1]['discharge'] = value
                recs[-1]['dchg_volt'] = voltage
            vstr = (' V=%dmV' % voltage) if voltage else ''
            self.on_log_message("RECORD[%#04x] discharge = %d mAh%s | records=%d" % (addr, value, vstr, len(recs)), "info")
            self._write_save_line(module_idx, 'discharge', value, voltage)


    def _touch_state_progress(self, state):
        state.last_progress_time = time.time()

    def _reset_state_timer(self, state, now=None):
        if now is None:
            now = time.time()
        state.stage_start_time = now
        state.last_progress_time = now
        state.last_retry_time = 0

    def _is_module_fresh(self, idx, max_age=TEST_START_FRESH_SEC):
        last_seen = self.module_last_seen[idx] if 0 <= idx < MODULE_COUNT else 0
        if not last_seen:
            return False
        mod = self.modules_data[idx] if 0 <= idx < MODULE_COUNT else None
        if not mod or not mod.get('executor_online', True):
            return False
        return (time.time() - last_seen) <= max_age

    def _fresh_online_modules(self):
        return sorted(idx for idx in self.online_modules if self._is_module_fresh(idx))

    def send_command_reliable(self, cmd, addr, data=None):
        """Non-blocking send: fire-and-forget + optional retry via timer."""
        if self.serial_thread is None or self.serial_thread.ser is None or not self.serial_thread.ser.is_open:
            return False
        self.serial_thread.send_command(cmd, addr, data)
        return True

    def send_command_confirmed(self, cmd, addr, data=None, retries=3):
        """Send with ACK confirmation (BLOCKING). Use sparingly."""
        return self._send_command_with_ack(cmd, addr, data, retries)

    def _send_command_with_ack(self, cmd, addr, data=None, retries=3):
        """Send command with ACK confirmation. Broadcast addr=0x00 fire-and-forget."""
        for attempt in range(retries):
            if self.serial_thread is None or self.serial_thread.ser is None or not self.serial_thread.ser.is_open:
                time.sleep(0.05)
                continue
            if addr != 0x00:
                status = self.serial_thread.send_command_sync(cmd, addr, data)
                if status == 0:
                    return True
                if status < 0:
                    self.on_log_message("RETRY %d/%d: CMD 0x%02X addr=%d TIMEOUT" % (attempt+1, retries, cmd, addr), "error")
                else:
                    self.on_log_message("RETRY %d/%d: CMD 0x%02X addr=%d FAIL status=0x%02X" % (attempt+1, retries, cmd, addr, status), "error")
                time.sleep(0.1)
            else:
                self.serial_thread.send_command(cmd, addr, data)
                time.sleep(0.02)
                return True
        self.on_log_message("GAVE UP: CMD 0x%02X addr=%d after %d retries" % (cmd, addr, retries), "error")
        return False

    def abort_module_test(self, idx, state, reason):
        addr = MODULE_ADDRS[idx]
        self.send_command_reliable(CMD_STOP_CHARGE, addr)
        self.send_command_reliable(CMD_STOP_DISCHARGE, addr)
        state.completed = True
        state.result = reason
        self.global_test_results[idx] = {
            'result': reason,
            'discharge': getattr(state, 'discharge_cap', 0),
            'charge2': getattr(state, 'charge2_cap', 0)
        }
        self._write_save_line(idx, 'fail_' + str(reason).replace(',', ';'), 0)
        self.check_all_completed()
        self.on_log_message(tr('abort_test_reason', addr, reason), "error")

    def check_all_completed(self):
        if self.loop_test_active:
            return
        if not self.global_test_mode:
            return
        if all(st.completed for st in self.module_test_states.values()):
            self.on_log_message(tr('all_modules_completed'), "info")
            if self.loop_test_active:
                self.stop_loop_test(final_stop=True)
            else:
                has_ng = any(self.global_test_results[i] and
                           self.global_test_results[i].get('result') == 'NG'
                           for i in range(MODULE_COUNT))
                self.stop_test(reset_ui=True)
                if has_ng and self.serial_thread:
                    self.serial_thread.send_all_command(CMD_ALL_BUZZER_ON)
                    self.buzzer_timer.start(30000)
                QMessageBox.information(self, tr('information'), tr('global_test_complete_msg'))


    def check_ng_and_buzz(self):
        if not self.serial_thread or not self.serial_thread.ser or not self.serial_thread.ser.is_open:
            return
        for idx in range(MODULE_COUNT):
            result = self.global_test_results[idx]
            if result and result.get('result') == 'NG':
                if idx in self.online_modules:
                    self.serial_thread.send_command(CMD_BUZZER_ON, MODULE_ADDRS[idx])
                    self.ng_buzzer_set.add(idx)

    def check_remaining_current(self):
        if not self.serial_thread or not self.serial_thread.ser or not self.serial_thread.ser.is_open:
            return
        for idx, state in self.module_test_states.items():
            if state.completed:
                mod = self.modules_data[idx]
                if mod:
                    chg_cur = mod.get('charge_current', 0)
                    dchg_cur = mod.get('discharge_current', 0)
                    if chg_cur > CURRENT_THRESHOLD or dchg_cur > CURRENT_THRESHOLD:
                        self.serial_thread.send_command(CMD_STOP_CHARGE, MODULE_ADDRS[idx])
                        self.serial_thread.send_command(CMD_STOP_DISCHARGE, MODULE_ADDRS[idx])

    def refresh_serial_ports(self):
        ports = [port.device for port in serial.tools.list_ports.comports()]
        self.port_combo.clear()
        if ports:
            self.port_combo.addItems(ports)
        else:
            self.port_combo.addItem(tr('no_ports_found'))

    # ---------- 串口连接 ----------
    def _load_logo(self):
        """Load company logo (supports PyInstaller bundled and dev paths)."""
        import os, sys
        from PyQt5.QtGui import QPixmap
        if getattr(sys, "frozen", False):
            base = sys._MEIPASS
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        logo_path = os.path.join(base, "logo.png")
        if os.path.exists(logo_path):
            pix = QPixmap(logo_path)
            if not pix.isNull():
                self.logo_label.setPixmap(pix.scaled(
                    self.logo_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))
                return
        self.logo_label.setText("LOGO")

    def on_connect(self):
        if self.serial_thread is None:
            port = self.port_combo.currentText()
            if port == tr('no_ports_found'):
                QMessageBox.critical(self, tr('error'), tr('no_ports_found'))
                return
            baud = int(self.baud_combo.currentText())
            self.serial_thread = ProtocolThread(port, baud)
            self.serial_thread.data_received.connect(self.on_data_received)
            self.serial_thread.error_occurred.connect(self.on_serial_error)
            self.serial_thread.log_message.connect(self.on_log_message)
            self.serial_thread.command_response.connect(self.on_command_response)
            self.serial_thread.first_scan_done.connect(self.send_default_commands)
            self.serial_thread.start()
            self.btn_connect.setText(tr('disconnect'))
            self.online_modules.clear()
            self.batch_read_received = False
            self._reset_display()
            self._peak_chg_mah = [0] * MODULE_COUNT
            self._peak_dchg_mah = [0] * MODULE_COUNT
            # wait for first_scan_done signal before sending defaults
        else:
            self.test_stage_timer.stop()
            self.buzzer_timer.stop()
            self.current_check_timer.stop()
            self.serial_thread.stop()
            self.serial_thread = None
            self._reset_display()
            self._peak_chg_mah = [0] * MODULE_COUNT
            self._peak_dchg_mah = [0] * MODULE_COUNT
            self.btn_connect.setText(tr('connect'))
            self.on_log_message(tr('disconnect'), "info")

    def send_default_commands(self):
        if self.serial_thread is None:
            return
        if not self.online_modules:
            self.on_log_message(tr('no_online_modules'), "info")
            return
        bat_type = 0x00
        cycle_num = 5
        self.serial_thread.send_all_command(CMD_ALL_SET_BAT_TYPE, [bat_type])
        QTimer.singleShot(20, lambda: self.serial_thread.send_all_command(CMD_ALL_SET_CYCLE_NUM, [cycle_num]))
        for idx in self.online_modules:
            self.set_cycle_nums[idx] = cycle_num
        self.on_log_message(tr('default_settings_sent'), "info")

    def on_command_response(self, cmd, status, data):
        cmd_name = CMD_NAMES.get(cmd, f"0x{cmd:02X}")
        if status == 0:
            msg = tr('command_succeeded', cmd_name)
        else:
            msg = tr('command_failed', cmd_name)
        self.statusBar.showMessage(msg, 2000)
        self.on_log_message(msg, "info")

    def on_emergency_stop(self):
        if not self.serial_thread:
            QMessageBox.warning(self, tr('warning'), tr('not_connected'))
            return
        reply = QMessageBox.question(self, tr('confirm'), tr('confirm_emergency_stop'),
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self._stop_auto_save()
            self.serial_thread.send_all_command(CMD_ALL_EMERGENCY_STOP)
            self._reset_display()
            self._peak_chg_mah = [0] * MODULE_COUNT
            self._peak_dchg_mah = [0] * MODULE_COUNT
            self.statusBar.showMessage(tr('emergency_stop_sent'), 3000)
            self.on_log_message(tr('emergency_stop_sent'), "info")
            self.stop_test(reset_ui=True)

    def on_stop_buzzer(self):
        self.buzzer_timer.stop()
        self.ng_buzzer_set.clear()
        if self.serial_thread and self.serial_thread.ser and self.serial_thread.ser.is_open:
            self.serial_thread.send_all_command(CMD_ALL_BUZZER_OFF)
            self.on_log_message(tr('buzzers_off'), "info")
            self.statusBar.showMessage(tr('buzzers_off'), 2000)

    def on_all_command(self, cmd):
        mapping = {
            CMD_START_CHARGE: CMD_ALL_START_CHARGE,
            CMD_STOP_CHARGE: CMD_ALL_STOP_CHARGE,
            CMD_START_DISCHARGE: CMD_ALL_START_DISCHARGE,
            CMD_STOP_DISCHARGE: CMD_ALL_STOP_DISCHARGE,
        }
        self.serial_thread.send_all_command(mapping[cmd])

    def on_set_cycle_num(self):
        idx = self.module_combo.currentIndex()
        addr = MODULE_ADDRS[idx]
        num = self.cycle_spin.value()
        if self.serial_thread:
            self.serial_thread.send_command(CMD_SET_CYCLE_NUM, addr, [num])
            self.set_cycle_nums[idx] = num
            self.on_log_message(tr('set_cycle_num_module', addr, num), "info")

    def on_set_cycle_num_all(self):
        num = self.cycle_spin.value()
        if self.serial_thread:
            self.serial_thread.send_all_command(CMD_ALL_SET_CYCLE_NUM, [num])
            for i in range(MODULE_COUNT):
                self.set_cycle_nums[i] = num
            self.on_log_message(tr('set_cycle_num_all', num), "info")

    def on_set_battery_type_all(self):
        if not self.serial_thread:
            QMessageBox.warning(self, tr('warning'), tr('not_connected'))
            return
        bat_type = self.battery_type_combo.currentData()
        if not self.online_modules:
            QMessageBox.information(self, tr('information'), tr('no_online_modules'))
            return
        reply = QMessageBox.question(self, tr('confirm'),
                                     tr('confirm_set_bat_type', bat_type),
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.serial_thread.send_all_command(CMD_ALL_SET_BAT_TYPE, [bat_type])
            self.on_log_message(tr('set_bat_type_all'), "info")

    def on_export_excel(self):
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.critical(self, tr('missing_library'), tr('install_openpyxl'))
            return

        # 自动生成文件名: 电池测试_YYYYMMDD_HHMMSS.xlsx
        from datetime import datetime
        default_name = f"电池测试_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, tr('export_excel'), default_name, "Excel Files (*.xlsx)")
        if not file_path:
            return

        wb = openpyxl.Workbook()
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        def style_header(ws, row_num, max_col):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row_num, column=c)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

        def style_cells(ws, start_row, end_row, max_col):
            if end_row is None or end_row < start_row:
                return
            for r in range(start_row, end_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.alignment = center_align
                    cell.border = thin_border

        def auto_width(ws):
            try:
                for col in ws.columns:
                    try:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            try:
                                v = str(cell.value) if cell.value is not None else ""
                                if len(v) > max_len:
                                    max_len = len(v)
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_len + 3, 32)
                    except:
                        pass
            except:
                pass

        # ===== Sheet 1: 实时数据 =====
        ws1 = wb.active
        ws1.title = "实时数据"
        r1_headers = ["模块", "地址", "电压(mV)", "充电电流(mA)", "放电电流(mA)",
                       "充电mAh", "放电mAh", "循环", "状态"]
        ws1.append(r1_headers)
        for i in range(MODULE_COUNT):
            mod = self._display[i]
            if mod is None:
                self.main_table.setItem(i, 0, QTableWidgetItem(f"{MODULE_ADDRS[i]}"))
                for c in range(1, self.main_table.columnCount()):
                    self.main_table.setItem(i, c, QTableWidgetItem(""))
                continue
            status_str = ""
            s = mod['status']
            if s & 0x40: status_str = "无电池"
            else:
                if s & 0x01: status_str += "充电 "
                if s & 0x02: status_str += "放电 "
                if s & 0x10: status_str += "充满 "
                if s & 0x20: status_str += "放完 "
                if s & 0x04: status_str += "过充! "
                if s & 0x08: status_str += "过放! "
                if not status_str: status_str = "空闲"
            ws1.append([
                f"模块{i+1}",
                f"{MODULE_ADDRS[i]}",
                mod['voltage'],
                mod['charge_current'],
                mod['discharge_current'],
                mod['charge_mah'],
                mod['discharge_mah'],
                mod['cycle'],
                status_str.strip()
            ])
        style_header(ws1, 1, len(r1_headers))
        style_cells(ws1, 2, ws1.max_row, len(r1_headers))
        auto_width(ws1)

        # ===== Sheet 2: 循环记录 =====
        ws2 = wb.create_sheet("循环记录")
        max_cycles = max((len(self.cycle_records[i]) for i in range(MODULE_COUNT)), default=0)
        r2_headers = ["模块", "地址"]
        for c in range(max_cycles):
            r2_headers.append(f"Cycle{c+1}_充电(mAh)")
            r2_headers.append(f"Cycle{c+1}_放电(mAh)")
        ws2.append(r2_headers)
        for i in range(MODULE_COUNT):
            row = [f"模块{i+1}", f"{MODULE_ADDRS[i]}"]
            records = self.cycle_records[i]
            for c in range(max_cycles):
                if c < len(records):
                    rec = records[c]
                    row.append(rec.get('charge') if rec.get('charge') is not None else "")
                    row.append(rec.get('discharge') if rec.get('discharge') is not None else "")
                else:
                    row.extend(["", ""])
            ws2.append(row)
        style_header(ws2, 1, len(r2_headers))
        style_cells(ws2, 2, ws2.max_row, len(r2_headers))
        auto_width(ws2)

        # ===== Sheet 3: 测试结果 =====
        ws3 = wb.create_sheet("测试结果")
        r3_headers = ["模块", "地址", "结果", "放电容量(mAh)", "充电2容量(mAh)",
                       "充电1最大(mAh)", "下限(mAh)", "上限(mAh)"]
        ws3.append(r3_headers)
        for i in range(MODULE_COUNT):
            result = self.global_test_results[i]
            if not result:
                continue
            state = self.module_test_states.get(i)
            lower = state.lower_limit if state else self.lower_limit_spin.value()
            upper = state.upper_limit if state else self.upper_limit_spin.value()
            chg1_max = state.charge1_max if state and hasattr(state, 'charge1_max') else ""
            ws3.append([
                f"模块{i+1}",
                f"{MODULE_ADDRS[i]}",
                result.get('result', ''),
                result.get('discharge', ''),
                result.get('charge2', ''),
                chg1_max,
                lower, upper
            ])
        style_header(ws3, 1, len(r3_headers))
        style_cells(ws3, 2, ws3.max_row, len(r3_headers))
        auto_width(ws3)

        try:
            wb.save(file_path)
            QMessageBox.information(self, tr('export_complete'), tr('export_path', file_path))
        except Exception as e:
            import traceback
            QMessageBox.critical(self, "导出失败", f"保存文件时出错:\n{str(e)}\n\n{traceback.format_exc()}")

    def show_log_window(self):
        if self.log_window is None:
            self.log_window = LogWindow(self)
            self.log_window.closed.connect(self.on_log_window_closed)
            if self.serial_thread:
                self.serial_thread.log_message.disconnect(self.on_log_message)
                self.serial_thread.log_message.connect(lambda m, c: self.log_window.append_log(m, c))
            self.log_window.show()
        else:
            self.log_window.raise_()
            self.log_window.activateWindow()

    def on_log_window_closed(self):
        self.log_window = None
        if self.serial_thread:
            try:
                self.serial_thread.log_message.disconnect()
            except TypeError:
                pass
            self.serial_thread.log_message.connect(self.on_log_message)

    # ---------- 数据处理 ----------
    def on_data_received(self, modules):
        now = time.time()
        for i, mod in enumerate(modules):
            if mod is None:
                self.module_miss_count[i] += 1
            else:
                self.module_last_seen[i] = now
                self.module_miss_count[i] = 0

        for mod in modules:
            if mod is None:
                continue
            addr = mod.get('addr')
            if addr not in MODULE_ADDRS:
                self.on_log_message("DROP invalid module addr: %s" % addr, "error")
                continue
            idx = MODULE_ADDRS.index(addr)
            self.modules_data[idx] = mod

            curr_status = mod['status']
            curr_charge_mah = mod['charge_mah']
            curr_discharge_mah = mod['discharge_mah']

            if (curr_status & 0x04) and not (self.last_status[idx] & 0x04):
                self.on_alarm(tr('alarm_overcharge', MODULE_ADDRS[idx]))
            if (curr_status & 0x08) and not (self.last_status[idx] & 0x08):
                self.on_alarm(tr('alarm_overdischarge', MODULE_ADDRS[idx]))

            if (curr_status & 0x01) and not (self.last_status[idx] & 0x01):
                self.on_log_message(tr('charge_started', MODULE_ADDRS[idx]), "info")
            if not (curr_status & 0x01) and (self.last_status[idx] & 0x01):
                self.on_log_message(tr('charge_stopped', MODULE_ADDRS[idx]), "info")
            if (curr_status & 0x02) and not (self.last_status[idx] & 0x02):
                self.on_log_message(tr('discharge_started', MODULE_ADDRS[idx]), "info")
            if not (curr_status & 0x02) and (self.last_status[idx] & 0x02):
                self.on_log_message(tr('discharge_stopped', MODULE_ADDRS[idx]), "info")

            self.last_status[idx] = curr_status
            self.last_charge_mah[idx] = curr_charge_mah
            self.last_discharge_mah[idx] = curr_discharge_mah
            self.last_cycle[idx] = mod['cycle']

        # Sync validated display buffer for all modules
        for i in range(MODULE_COUNT):
            self._sync_display(i)

        current_online = set()
        for i, mod in enumerate(modules):
            if mod is None:
                continue
            if (mod['voltage'] > 0 or mod['charge_current'] > 0 or
                mod['discharge_current'] > 0 or mod['charge_mah'] > 0 or
                mod['discharge_mah'] > 0 or mod['status'] != 0):
                current_online.add(i)
        self.online_modules = current_online
        self.batch_read_received = True

        # 每10批输出一次首个模块数据
        if not hasattr(self, '_data_log_cnt'): self._data_log_cnt = 0
        self._data_log_cnt += 1
        if self._data_log_cnt % 20 == 0 and current_online:
            i = min(current_online)
            m = self.modules_data[i]
            self.on_log_message(
                f"[{MODULE_ADDRS[i]}] V={m['voltage']}mV "
                f"Chg={m['charge_mah']}mAh/{m['charge_current']}mA "
                f"Dchg={m['discharge_mah']}mAh/{m['discharge_current']}mA "
                f"St=0x{m['status']:02X} Cyc={m['cycle']}",
                "comm")

        if self.loop_test_active:
            for idx in list(self.module_test_states.keys()):
                state = self.module_test_states.get(idx)
                if isinstance(state, LoopModuleState) and not state.completed:
                    mod = modules[idx] if 0 <= idx < len(modules) else None
                    try:
                        if mod:
                            if not mod.get('executor_online', True):
                                self._abort_state_as_failed(idx, state, "executor offline")
                            else:
                                self.update_loop_state(idx, state, mod)
                        else:
                            self._handle_stale_state(idx, state)
                    except Exception as e:
                        self.on_log_message("STATE ERROR[%#04x] %s" % (MODULE_ADDRS[idx], e), "error")
                        self._abort_state_as_failed(idx, state, "state exception")
        else:
            for idx in list(self.module_test_states.keys()):
                state = self.module_test_states.get(idx)
                if isinstance(state, ModuleTestState) and not state.completed:
                    mod = modules[idx] if 0 <= idx < len(modules) else None
                    try:
                        if mod:
                            if not mod.get('executor_online', True):
                                self._abort_state_as_failed(idx, state, "executor offline")
                            else:
                                self.update_module_test_state(idx, state, mod)
                        else:
                            self._handle_stale_state(idx, state)
                    except Exception as e:
                        self.on_log_message("STATE ERROR[%#04x] %s" % (MODULE_ADDRS[idx], e), "error")
                        self._abort_state_as_failed(idx, state, "state exception")

        self.update_table()

    def update_loop_state(self, idx, state, mod):
        if state.completed:
            return
        status = mod['status']
        addr = MODULE_ADDRS[idx]
        now = time.time()
        if self._check_state_deadline(idx, state, now):
            return

        # --- Alarm handling with retry ---
        # Overcharge during charging: stop and retry after cooldown
        if status & 0x04 and state.stage == STAGE_CHARGE1:
            if not getattr(state, 'alarm_oc_logged', False):
                state.alarm_oc_logged = True
                self.on_log_message("ALARM: %#04x overcharge | cycle %d/%d | V=%dmV" % (addr, state.current_cycle+1, state.total_cycles, mod.get("voltage",0)), "error")
            # Stop charge, retry start after 10s cooldown
            if now - state.stage_start_time > 10:
                self.send_command_reliable(CMD_STOP_CHARGE, addr)
                time.sleep(0.1)
                self.send_command_reliable(CMD_START_CHARGE, addr)
                state.charge_start = 0
                state.stage_start_time = now
                self.on_log_message("RETRY: %#04x restart charge (alarm cleared)" % addr, "error")
            return
        else:
            state.alarm_oc_logged = False

        # Overdischarge during discharging: stop and retry
        if status & 0x08 and state.stage == STAGE_DISCHARGE:
            if not getattr(state, 'alarm_od_logged', False):
                state.alarm_od_logged = True
                self.on_log_message("ALARM: %#04x overdischarge | cycle %d/%d | V=%dmV" % (addr, state.current_cycle+1, state.total_cycles, mod.get("voltage",0)), "error")
            if now - state.stage_start_time > 10:
                self.send_command_reliable(CMD_STOP_DISCHARGE, addr)
                time.sleep(0.1)
                self.send_command_reliable(CMD_START_DISCHARGE, addr)
                state.discharge_start = 0
                state.stage_start_time = now
                self.on_log_message("RETRY: %#04x restart discharge (alarm cleared)" % addr, "error")
                QTimer.singleShot(1500, lambda a=addr, i=idx: self._verify_discharge_started(a, i))
            return
        else:
            state.alarm_od_logged = False

        # No-battery alarm: log once, cannot retry
        if status & 0x40:
            if not getattr(state, 'alarm_nb_logged', False):
                state.alarm_nb_logged = True
                self.on_log_message("ALARM: %#04x no-battery | cycle %d/%d" % (addr, state.current_cycle+1, state.total_cycles), "error")
            self._abort_state_as_failed(idx, state, "no battery")
            return
        else:
            state.alarm_nb_logged = False

        # --- Discharge stall recovery: module in DISCHARGE stage but not discharging ---
        if state.stage == STAGE_DISCHARGE and not (status & 0x02) and not (status & 0x20):
            elapsed = now - state.stage_start_time
            if elapsed > 3.0:
                self.on_log_message("WARN: %#04x DISCHARGE stage but not discharging (%.1fs), retry cmd" % (addr, elapsed), "error")
                self.send_command_reliable(CMD_START_DISCHARGE, addr)
                state.discharge_start = mod.get('discharge_mah', 0) if mod else 0
                state.stage_start_time = now
            return

        # --- Plateau detection (fallback for missed status bits) ---
        PLATEAU_THRESH = 10
        if state.stage == STAGE_CHARGE1:
            if mod['charge_mah'] == state._last_chg:
                state._chg_stable += 1
            else:
                state._chg_stable = 0
            state._last_chg = mod['charge_mah']
        elif state.stage == STAGE_DISCHARGE:
            if mod['discharge_mah'] == state._last_dchg:
                state._dchg_stable += 1
            else:
                state._dchg_stable = 0
            state._last_dchg = mod['discharge_mah']

        # --- Normal stage transitions ---
        if state.stage == STAGE_CHARGE1:
            if status & 0x10:
                self.send_command_reliable(CMD_STOP_CHARGE, addr)
                time.sleep(0.05)
                chg_val = max(0, state.charge_max - state.charge_start)
                self.add_cycle_record(idx, 'charge', chg_val, mod['voltage'])
                state.stage = STAGE_DISCHARGE
                state.charge_end = mod['charge_mah']
                self.send_command_reliable(CMD_START_DISCHARGE, addr)
                state.discharge_start = 0
                self._reset_state_timer(state, now)
                self.on_log_message("CYCLE %d/%d: %#04x charge %dmAh -> discharge" % (state.current_cycle+1, state.total_cycles, addr, chg_val), "info")
                # Verify discharge actually started in 1.5s
                QTimer.singleShot(1500, lambda a=addr, i=idx: self._verify_discharge_started(a, i))
            elif state._chg_stable >= PLATEAU_THRESH and mod['charge_mah'] > 200 and now - state.stage_start_time > 30:
                # FALLBACK: charge_mah plateaued for 5 polls, charge likely complete but status bit missed
                self.send_command_reliable(CMD_STOP_CHARGE, addr)
                time.sleep(0.05)
                chg_val = max(0, state.charge_max - state.charge_start)
                self.add_cycle_record(idx, 'charge', chg_val, mod['voltage'])
                state.stage = STAGE_DISCHARGE
                state.charge_end = mod['charge_mah']
                self.send_command_reliable(CMD_START_DISCHARGE, addr)
                state.discharge_start = 0
                state._dchg_stable = 0
                self._reset_state_timer(state, now)
                self.on_log_message("PLATEAU FALLBACK[%#04x] charge %dmAh stable=%d -> discharge" % (addr, chg_val, state._chg_stable), "error")
                QTimer.singleShot(1500, lambda a=addr, i=idx: self._verify_discharge_started(a, i))
        elif state.stage == STAGE_DISCHARGE:
            if status & 0x20:
                self.send_command_reliable(CMD_STOP_DISCHARGE, addr)
                time.sleep(0.05)
                disc = max(0, state.discharge_max - state.discharge_start)
                self.add_cycle_record(idx, 'discharge', disc, mod['voltage'])
                state.current_cycle += 1

                if state.current_cycle >= state.total_cycles:
                    state.completed = True
                    self.on_log_message("DONE: %#04x all %d cycles completed" % (addr, state.total_cycles), "info")
                    self.check_loop_all_completed()
                else:
                    state.stage = STAGE_CHARGE1
                    self.send_command_reliable(CMD_START_CHARGE, addr)
                    state.charge_start = 0
                    state.charge_max = 0
                    self._reset_state_timer(state, now)
                    self.on_log_message("CYCLE %d/%d: %#04x discharge %dmAh -> charge" % (state.current_cycle, state.total_cycles, addr, disc), "info")
            elif state._dchg_stable >= PLATEAU_THRESH and mod['discharge_mah'] > 200 and now - state.stage_start_time > 30:
                # FALLBACK: discharge_mah plateaued, discharge likely complete but status bit missed
                self.send_command_reliable(CMD_STOP_DISCHARGE, addr)
                time.sleep(0.05)
                disc = max(0, state.discharge_max - state.discharge_start)
                self.add_cycle_record(idx, 'discharge', disc, mod['voltage'])
                state.current_cycle += 1
                state._chg_stable = 0
                if state.current_cycle >= state.total_cycles:
                    state.completed = True
                    self.on_log_message("DONE+PLATEAU: %#04x all %d cycles completed" % (addr, state.total_cycles), "info")
                    self.check_loop_all_completed()
                else:
                    state.stage = STAGE_CHARGE1
                    self.send_command_reliable(CMD_START_CHARGE, addr)
                    state.charge_start = 0
                    state.charge_max = 0
                    self._reset_state_timer(state, now)
                    self.on_log_message("PLATEAU FALLBACK[%#04x] discharge %dmAh stable=%d -> charge" % (addr, disc, state._dchg_stable), "error")

        if state.stage == STAGE_CHARGE1 and (status & 0x01):
            if mod['charge_mah'] > state.charge_max:
                state.charge_max = mod['charge_mah']
                self._touch_state_progress(state)
        elif state.stage == STAGE_DISCHARGE and (status & 0x02):
            if mod['discharge_mah'] > state.discharge_max:
                state.discharge_max = mod['discharge_mah']
                self._touch_state_progress(state)

    def _verify_discharge_started(self, addr, idx, attempt=1):
        """Check if discharge actually started after CMD_START_DISCHARGE. Retry if not."""
        if not self.test_running:
            return
        state = self.module_test_states.get(idx)
        if state is None or state.completed:
            return
        now = time.time()
        last_seen = self.module_last_seen[idx] or state.stage_start_time
        if now - last_seen >= DATA_STALE_RETRY_SEC:
            self._handle_stale_state(idx, state)
            return
        mod = self.modules_data[idx]
        if mod is None:
            self._handle_stale_state(idx, state)
            return
        status = mod['status']
        if status & 0x02:  # Discharging active
            return  # OK, discharge running
        if status & 0x20:  # Discharge already complete
            return  # OK, already done
        # Module NOT discharging. Retry with ACK confirmation.
        self.on_log_message("VERIFY FAIL[%#04x] not discharging (attempt %d), retry with ACK" % (addr, attempt), "error")
        if attempt <= 2:
            confirmed = self.serial_thread and self.serial_thread.send_command_sync(CMD_START_DISCHARGE, addr, None, 200)
            if confirmed == 0:
                self.on_log_message("VERIFY[%#04x] START_DISCHARGE ACK confirmed, waiting status" % addr, "info")
                state.discharge_start = 0
                state.last_retry_time = now
                QTimer.singleShot(1000, lambda: self._verify_discharge_started(addr, idx, attempt+1))
            else:
                self.on_log_message("VERIFY FAIL[%#04x] ACK timeout, will retry" % addr, "error")
                QTimer.singleShot(1000, lambda: self._verify_discharge_started(addr, idx, attempt+1))
        else:
            self.on_log_message("ABORT[%#04x] discharge cannot start after 3 attempts" % addr, "error")
            if isinstance(state, ModuleTestState):
                self.abort_module_test(idx, state, 'NG')
            else:
                state.completed = True
                self.check_loop_all_completed()

    def check_loop_all_completed(self):
        if not self.loop_test_active:
            return
        if all(state.completed for state in self.module_test_states.values()):
            self.on_log_message(tr('all_modules_completed_loop'), "info")
            self.on_loop_completed()

    def update_module_test_state(self, idx, state, mod):
        if state.completed:
            return
        addr = MODULE_ADDRS[idx]
        status = mod['status']
        lower = state.lower_limit
        upper = state.upper_limit
        now = time.time()
        if self._check_state_deadline(idx, state, now):
            return

        if status & 0x04:
            self.on_log_message(tr('alarm_overcharge', addr), "error")
            self.abort_module_test(idx, state, 'NG')
            return
        if status & 0x08:
            self.on_log_message(tr('alarm_overdischarge', addr), "error")
            self.abort_module_test(idx, state, 'NG')
            return
        if status & 0x40:
            self.on_log_message("ALARM: %#04x no-battery during test" % addr, "error")
            self.abort_module_test(idx, state, 'NG')
            return

        # --- Plateau detection ---
        PMT_THRESH = 10
        if state.stage == STAGE_CHARGE1:
            if mod['charge_mah'] == state._last_chg:
                state._chg_stable += 1
            else:
                state._chg_stable = 0
            state._last_chg = mod['charge_mah']
        elif state.stage == STAGE_DISCHARGE:
            if mod['discharge_mah'] == state._last_dchg:
                state._dchg_stable += 1
            else:
                state._dchg_stable = 0
            state._last_dchg = mod['discharge_mah']
        elif state.stage == STAGE_CHARGE2:
            if mod['charge_mah'] == state._last_chg:
                state._chg_stable += 1
            else:
                state._chg_stable = 0
            state._last_chg = mod['charge_mah']

        if state.stage == STAGE_CHARGE1:
            if status & 0x10:
                self.on_log_message(tr('charge1_completed', addr), "info")
                self.send_command_reliable(CMD_STOP_CHARGE, addr)
                time.sleep(0.05)
                chg_val = max(0, state.charge1_max - state.charge1_start)
                self.add_cycle_record(idx, 'charge', chg_val, mod['voltage'])
                state.stage = STAGE_DISCHARGE
                state.charge_end = mod['charge_mah']
                self.send_command_reliable(CMD_START_DISCHARGE, addr)
                state.discharge_start = 0  # STC8H resets on start
                state.discharge_max = 0  # STC8H resets on start
                self._reset_state_timer(state, now)
                self.on_log_message(tr('discharge_started', addr), "info")
                QTimer.singleShot(1500, lambda a=addr, i=idx: self._verify_discharge_started(a, i))
                return
            elif state._chg_stable >= PMT_THRESH and mod['charge_mah'] > 200 and now - state.stage_start_time > 30:
                self.on_log_message("PLATEAU FALLBACK: %#04x charge1 completed via plateau (%dmAh)" % (addr, mod['charge_mah']), "error")
                self.send_command_reliable(CMD_STOP_CHARGE, addr)
                time.sleep(0.05)
                chg_val = max(0, state.charge1_max - state.charge1_start)
                self.add_cycle_record(idx, 'charge', chg_val, mod['voltage'])
                state.stage = STAGE_DISCHARGE
                state.charge_end = mod['charge_mah']
                self.send_command_reliable(CMD_START_DISCHARGE, addr)
                state.discharge_start = 0
                state.discharge_max = 0
                state._dchg_stable = 0
                self._reset_state_timer(state, now)
                self.on_log_message(tr('discharge_started', addr), "info")
                QTimer.singleShot(1500, lambda a=addr, i=idx: self._verify_discharge_started(a, i))
                return
        elif state.stage == STAGE_DISCHARGE:
            if status & 0x20:
                self.on_log_message(tr('discharge_completed', addr), "info")
                self.send_command_reliable(CMD_STOP_DISCHARGE, addr)
                time.sleep(0.05)
                discharged = state.discharge_max - state.discharge_start
                state.discharge_cap = discharged
                self.add_cycle_record(idx, 'discharge', discharged, mod['voltage'])

                state.stage = STAGE_CHARGE2
                self._reset_state_timer(state, now)
                self.send_command_reliable(CMD_START_CHARGE, addr)
                state.charge2_start = 0  # STC8H resets on start
                state.charge2_max = 0  # STC8H resets on start
                state.charge2_empty_retries = 0
                self.on_log_message(tr('discharge_done', addr, discharged), "info")
                return
            elif state._dchg_stable >= PMT_THRESH and mod['discharge_mah'] > 200 and now - state.stage_start_time > 30:
                self.on_log_message("PLATEAU FALLBACK: %#04x discharge completed via plateau (%dmAh)" % (addr, mod['discharge_mah']), "error")
                self.send_command_reliable(CMD_STOP_DISCHARGE, addr)
                time.sleep(0.05)
                discharged = state.discharge_max - state.discharge_start
                state.discharge_cap = discharged
                self.add_cycle_record(idx, 'discharge', discharged, mod['voltage'])
                state.stage = STAGE_CHARGE2
                self._reset_state_timer(state, now)
                self.send_command_reliable(CMD_START_CHARGE, addr)
                state.charge2_start = 0
                state.charge2_max = 0
                state.charge2_empty_retries = 0
                state._chg_stable = 0
                self.on_log_message(tr('discharge_done', addr, discharged), "info")
                return
        elif state.stage == STAGE_CHARGE2:
            if status & 0x10:
                self.on_log_message(tr('charge2_completed', addr), "info")
                self.send_command_reliable(CMD_STOP_CHARGE, addr)
                time.sleep(0.05)
                charged = state.charge2_max - state.charge2_start
                if charged < 1:
                    retries = getattr(state, 'charge2_empty_retries', 0) + 1
                    state.charge2_empty_retries = retries
                    self.on_log_message(tr('charge2_empty_retry', addr), "error")
                    if retries >= 3:
                        self.abort_module_test(idx, state, 'NG')
                        return
                    self.send_command_reliable(CMD_START_CHARGE, addr)
                    return
                state.charge2_cap = charged
                self.add_cycle_record(idx, 'charge', charged, mod['voltage'])
                ok = (lower <= state.discharge_cap <= upper) and (lower <= state.charge2_cap <= upper)
                result = "OK" if ok else "NG"
                self.global_test_results[idx] = {
                    'result': result,
                    'discharge': state.discharge_cap,
                    'charge2': state.charge2_cap
                }
                state.result = result
                state.stage = STAGE_COMPLETED
                state.completed = True
                self.send_command_reliable(CMD_STOP_CHARGE, addr)
                self.send_command_reliable(CMD_STOP_DISCHARGE, addr)
                self.on_log_message(tr('test_finished_result', addr, result), "info")
                self.check_all_completed()
                return
            elif state._chg_stable >= PMT_THRESH and mod['charge_mah'] > 200 and now - state.stage_start_time > 30:
                self.on_log_message("PLATEAU FALLBACK: %#04x charge2 completed via plateau (%dmAh)" % (addr, mod['charge_mah']), "error")
                self.send_command_reliable(CMD_STOP_CHARGE, addr)
                time.sleep(0.05)
                charged = state.charge2_max - state.charge2_start
                if charged < 1:
                    charged = mod['charge_mah']  # fallback: use current value
                state.charge2_cap = charged
                self.add_cycle_record(idx, 'charge', charged, mod['voltage'])
                ok = (lower <= state.discharge_cap <= upper) and (lower <= state.charge2_cap <= upper)
                result = "OK" if ok else "NG"
                self.global_test_results[idx] = {'result': result, 'discharge': state.discharge_cap, 'charge2': state.charge2_cap}
                state.result = result
                state.stage = STAGE_COMPLETED
                state.completed = True
                self.send_command_reliable(CMD_STOP_CHARGE, addr)
                self.send_command_reliable(CMD_STOP_DISCHARGE, addr)
                self.on_log_message(tr('test_finished_result', addr, result), "info")
                self.check_all_completed()
                return

        if state.stage == STAGE_CHARGE1 and (status & 0x01):
            if mod['charge_mah'] > state.charge1_max:
                state.charge1_max = mod['charge_mah']
                self._touch_state_progress(state)
        elif state.stage == STAGE_DISCHARGE and (status & 0x02):
            if mod['discharge_mah'] > state.discharge_max:
                state.discharge_max = mod['discharge_mah']
                self._touch_state_progress(state)
        elif state.stage == STAGE_CHARGE2 and (status & 0x01):
            if mod['charge_mah'] > state.charge2_max:
                state.charge2_max = mod['charge_mah']
                self._touch_state_progress(state)

    def on_test_timeout(self):
        if not self.test_running:
            return
        self.on_log_message(tr('test_timeout'), "error")
        self.stop_test(reset_ui=True)
        QMessageBox.critical(self, tr('error'), tr('test_timeout'))

    def stop_test(self, reset_ui=True):
        if not self.test_running and reset_ui:
            return
        self.test_running = False
        self.global_test_mode = False
        self.test_stage_timer.stop()
        self.buzzer_timer.stop()
        self.current_check_timer.stop()
        self.watchdog_timer.stop()
        self._stop_auto_save()
        if not self.test_running:
            self.history_timer.start(500)

        if self.serial_thread and self.serial_thread.ser and self.serial_thread.ser.is_open:
            self.serial_thread.send_all_command(CMD_ALL_BUZZER_OFF)
            self.serial_thread.send_all_command(CMD_ALL_EMERGENCY_STOP)
            self._peak_chg_mah = [0] * MODULE_COUNT
            self._peak_dchg_mah = [0] * MODULE_COUNT
            self.on_log_message(tr('stop_all_modules'), "info")

        if reset_ui:
            self.test_status_label.setText(tr('test_status'))
            self.global_test_status_label.setText(tr('global_test_status_idle'))
            self.btn_start_test.setText(tr('start_test_on_selected'))
            self.btn_stop_test.setEnabled(False)
            self.set_controls_enabled(True)
            self.btn_global_test.setEnabled(True)
            self.btn_start_test.setEnabled(True)
            self.btn_loop_test.setEnabled(True)
            self.btn_stop_loop.setEnabled(False)

        self.module_test_states.clear()
        self.single_test_state = None
        self.update_table()

    def set_controls_enabled(self, enabled):
        self.btn_start_charge.setEnabled(enabled)
        self.btn_stop_charge.setEnabled(enabled)
        self.btn_start_discharge.setEnabled(enabled)
        self.btn_stop_discharge.setEnabled(enabled)
        self.btn_all_start_charge.setEnabled(enabled)
        self.btn_all_stop_charge.setEnabled(enabled)
        self.btn_all_start_discharge.setEnabled(enabled)
        self.btn_all_stop_discharge.setEnabled(enabled)
        self.btn_set_cycle.setEnabled(enabled)
        self.btn_set_cycle_all.setEnabled(enabled)
        self.btn_set_bat_type_all.setEnabled(enabled)
        self.btn_export_excel.setEnabled(enabled)
        self.btn_stop_buzzer.setEnabled(True)

    def on_start_global_test(self):
        if not self.serial_thread:
            QMessageBox.warning(self, tr('warning'), tr('not_connected'))
            return
        if self.test_running:
            QMessageBox.warning(self, tr('warning'), tr('test_running'))
            return
        fresh_modules = self._fresh_online_modules()
        if not fresh_modules:
            QMessageBox.warning(self, tr('warning'), tr('no_online_modules'))
            return
        self.cycle_records = [[] for _ in range(MODULE_COUNT)]
        self.global_test_results = [None] * MODULE_COUNT
        self._reset_display()
        self.update_table()
        self._peak_chg_mah = [0] * MODULE_COUNT
        self._peak_dchg_mah = [0] * MODULE_COUNT
        self.update_history_table()
        lower = self.lower_limit_spin.value()
        upper = self.upper_limit_spin.value()
        reply = QMessageBox.question(self, tr('confirm'),
                                     tr('confirm_start_global_test', len(fresh_modules), lower, upper),
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        self.global_test_mode = True
        self.test_running = True
        self.module_test_states.clear()
        self.global_test_results = [None] * MODULE_COUNT

        self._global_test_modules = fresh_modules
        for idx in self._global_test_modules:
            mod = self.modules_data[idx]
            if mod is None:
                continue
            state = ModuleTestState(idx, lower, upper)
            state.start(mod['charge_mah'], mod['discharge_mah'])
            self.module_test_states[idx] = state
        if not self.module_test_states:
            self.test_running = False
            self.global_test_mode = False
            QMessageBox.warning(self, tr('warning'), tr('no_online_modules'))
            return

        if not self._start_auto_save():
            self.stop_test(reset_ui=True)
            return
        self.serial_thread.send_all_command(CMD_ALL_EMERGENCY_STOP)
        QTimer.singleShot(200, lambda: self.serial_thread.send_all_command(CMD_ALL_START_CHARGE))

        self.set_controls_enabled(False)
        self.btn_stop_test.setEnabled(True)
        self.btn_global_test.setEnabled(False)
        self.btn_start_test.setEnabled(False)
        self.btn_loop_test.setEnabled(False)

        self.global_test_status_label.setText(tr('global_test_independent'))
        self.test_status_label.setText(tr('test_running_global'))
        self.watchdog_timer.start(3000)
        self.on_log_message(tr('global_test_started', len(self.module_test_states)), "info")

    def on_start_test(self):
        self.cycle_records = [[] for _ in range(MODULE_COUNT)]
        self.global_test_results = [None] * MODULE_COUNT
        self._reset_display()
        self._peak_chg_mah = [0] * MODULE_COUNT
        self._peak_dchg_mah = [0] * MODULE_COUNT
        self.update_table()
        self.update_history_table()
        if not self.serial_thread:
            QMessageBox.warning(self, tr('warning'), tr('not_connected'))
            return
        if self.test_running:
            reply = QMessageBox.question(self, tr('confirm'), tr('test_running') + "\n" + tr('confirm_stop_test'),
                                         QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.stop_test(reset_ui=True)
            return

        idx = self.module_combo.currentIndex()
        addr = MODULE_ADDRS[idx]
        if self.modules_data[idx] is None or not self._is_module_fresh(idx):
            QMessageBox.warning(self, tr('warning'), tr('module_no_data', addr))
            return

        lower = self.lower_limit_spin.value()
        upper = self.upper_limit_spin.value()
        reply = QMessageBox.question(self, tr('confirm'), tr('confirm_start_test', addr),
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        self.global_test_mode = True
        self.test_running = True
        self.module_test_states.clear()
        self.global_test_results = [None] * MODULE_COUNT

        mod = self.modules_data[idx]
        state = ModuleTestState(idx, lower, upper)
        state.start(mod['charge_mah'], mod['discharge_mah'])
        self.module_test_states[idx] = state

        if not self._start_auto_save():
            self.stop_test(reset_ui=True)
            return
        self.serial_thread.send_command(CMD_STOP_CHARGE, addr)
        self.serial_thread.send_command(CMD_STOP_DISCHARGE, addr)
        QTimer.singleShot(200, lambda: self.serial_thread.send_command(CMD_START_CHARGE, addr))

        self.test_status_label.setText(tr('single_test_first_charge', addr))
        self.watchdog_timer.start(3000)
        self.set_controls_enabled(False)
        self.btn_stop_test.setEnabled(True)
        self.btn_start_test.setText(tr('stop_test'))
        self.btn_global_test.setEnabled(False)
        self.btn_loop_test.setEnabled(False)
        self.on_log_message(tr('single_test_started', addr), "info")

    def on_stop_test_clicked(self):
        if self.test_running:
            reply = QMessageBox.question(self, tr('confirm'), tr('confirm_stop_test'),
                                         QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.stop_test(reset_ui=True)

    def on_start_loop_test(self):
        if not self.serial_thread:
            QMessageBox.warning(self, tr('warning'), tr('not_connected'))
            return
        if self.test_running:
            QMessageBox.warning(self, tr('warning'), tr('test_running'))
            return
        fresh_modules = self._fresh_online_modules()
        if not fresh_modules:
            QMessageBox.warning(self, tr('warning'), tr('no_online_modules'))
            return

        total = self.cycle_spin.value()
        if total < 1:
            QMessageBox.warning(self, tr('warning'), tr('cycle_num_warning'))
            return

        reply = QMessageBox.question(self, tr('confirm'),
                                     tr('confirm_start_loop', len(fresh_modules), total),
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        self.loop_test_active = True
        self.test_running = True
        self.global_test_mode = True
        self.loop_total_cycles = total
        self.loop_current_cycle = 0

        self.cycle_records = [[] for _ in range(MODULE_COUNT)]
        self.global_test_results = [None] * MODULE_COUNT
        self._reset_display()
        self.update_history_table()
        self._peak_chg_mah = [0] * MODULE_COUNT
        self._peak_dchg_mah = [0] * MODULE_COUNT

        # Snapshot online modules at loop start
        self._loop_start_modules = fresh_modules
        self.module_test_states.clear()
        for idx in self._loop_start_modules:
            mod = self.modules_data[idx]
            if mod is None:
                continue
            state = LoopModuleState(idx, total)
            state.start(mod['charge_mah'], mod['discharge_mah'])
            self.module_test_states[idx] = state
        if not self.module_test_states:
            self.loop_test_active = False
            self.test_running = False
            self.global_test_mode = False
            self._stop_auto_save()
            QMessageBox.warning(self, tr('warning'), tr('no_online_modules'))
            return
        if not self._start_auto_save():
            self.stop_loop_test(final_stop=False)
            return

        self.serial_thread.send_all_command(CMD_ALL_EMERGENCY_STOP)
        QTimer.singleShot(200, lambda: self.serial_thread.send_all_command(CMD_ALL_START_CHARGE))

        self.set_controls_enabled(False)
        self.btn_global_test.setEnabled(False)
        self.btn_start_test.setEnabled(False)
        self.btn_loop_test.setEnabled(False)
        self.btn_stop_loop.setEnabled(True)
        self.btn_stop_test.setEnabled(False)

        self.global_test_status_label.setText(tr('loop_test_status', 0, total))
        self.test_status_label.setText(tr('loop_test_running'))
        self.watchdog_timer.start(3000)
        self.on_log_message(tr('loop_test_started', total), "info")

    def on_stop_loop_test(self):
        if self.loop_test_active:
            reply = QMessageBox.question(self, tr('confirm'), tr('confirm_stop_loop'),
                                         QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.stop_loop_test(final_stop=False)

    def on_loop_completed(self):
        if not self.loop_test_active:
            return
        self.update_history_table()
        self.on_log_message(tr('loop_completed'), "info")
        self.serial_thread.send_all_command(CMD_ALL_EMERGENCY_STOP)
        self._peak_chg_mah = [0] * MODULE_COUNT
        self._peak_dchg_mah = [0] * MODULE_COUNT
        QTimer.singleShot(500, lambda: self.stop_loop_test(final_stop=True))
        self.global_test_status_label.setText(tr('loop_test_status', self.loop_total_cycles, self.loop_total_cycles))

    def start_next_cycle(self):
        if not self.loop_test_active:
            return
        self.loop_current_cycle += 1
        self.global_test_status_label.setText(tr('loop_test_status', self.loop_current_cycle, self.loop_total_cycles))
        self.on_log_message(f"Starting loop cycle {self.loop_current_cycle}/{self.loop_total_cycles}", "info")

        # Snapshot online modules at cycle start to avoid mid-cycle drops
        cycle_modules = self._fresh_online_modules()
        if not cycle_modules:
            self.on_log_message("WARNING: no online modules at cycle start, aborting loop", "error")
            self.stop_loop_test(final_stop=True)
            return
        self.module_test_states.clear()
        for idx in cycle_modules:
            mod = self.modules_data[idx]
            if mod is None:
                continue
            state = LoopModuleState(idx, self.loop_total_cycles)
            state.start(mod['charge_mah'], mod['discharge_mah'])
            self.module_test_states[idx] = state

        self.serial_thread.send_all_command(CMD_ALL_EMERGENCY_STOP)
        QTimer.singleShot(200, lambda: self.serial_thread.send_all_command(CMD_ALL_START_CHARGE))

    def stop_loop_test(self, final_stop=False):
        if not self.loop_test_active and not final_stop:
            return

        self.loop_test_active = False
        self.test_running = False
        self.global_test_mode = False

        if self.serial_thread and self.serial_thread.ser and self.serial_thread.ser.is_open:
            self.serial_thread.send_all_command(CMD_ALL_EMERGENCY_STOP)
            self._peak_chg_mah = [0] * MODULE_COUNT
            self._peak_dchg_mah = [0] * MODULE_COUNT
            self.serial_thread.send_all_command(CMD_ALL_BUZZER_OFF)

        self.set_controls_enabled(True)
        self.btn_global_test.setEnabled(True)
        self.btn_start_test.setEnabled(True)
        self.btn_loop_test.setEnabled(True)
        self.btn_stop_loop.setEnabled(False)
        self.global_test_status_label.setText(tr('global_test_status_idle'))
        self.test_status_label.setText(tr('test_status'))
        self.test_stage_timer.stop()

        self.watchdog_timer.stop()
        self._stop_auto_save()
        self.module_test_states.clear()
        self.update_table()
        self.update_history_table()

        if final_stop:
            QMessageBox.information(self, tr('information'), tr('loop_completed_msg', self.loop_total_cycles))
        else:
            QMessageBox.information(self, tr('information'), tr('loop_stopped_user'))


    def _make_display_record(self, raw):
        """Create a display-safe record from raw module data."""
        return {
            'voltage': raw['voltage'],
            'charge_current': raw['charge_current'],
            'discharge_current': raw['discharge_current'],
            'charge_mah': raw['charge_mah'],
            'discharge_mah': raw['discharge_mah'],
            'status': raw['status'],
            'cycle': raw['cycle'],
            'ck_fail_count': raw.get('ck_fail_count', 0),
            'bridge_online': raw.get('bridge_online', True),
            'executor_online': raw.get('executor_online', True),
            'has_reading': raw['voltage'] > 0 or raw['charge_current'] > 0 or
                          raw['discharge_current'] > 0 or raw['status'] != 0
        }

    def _sync_display(self, idx):
        """Validate raw mod data with hysteresis: 3 valid to show, 3 invalid to hide."""
        raw = self.modules_data[idx]
        if raw is None:
            self._nobat_debounce[idx] = 0
            self._display[idx] = None
            return
        if raw.get("bridge_online") and not raw.get("executor_online", True):
            self._nobat_debounce[idx] = 0
            self._display[idx] = self._make_display_record(raw)
            return
        is_nobat = (raw["status"] & 0x40) or (raw["voltage"] == 0 and raw["charge_current"] == 0 and raw["discharge_current"] == 0)
        disp = self._display[idx]
        if disp is None:
            if is_nobat:
                self._nobat_debounce[idx] = 0
                return
            self._nobat_debounce[idx] += 1
            if self._nobat_debounce[idx] < 3:
                return
            self._nobat_debounce[idx] = 3
            self._display[idx] = self._make_display_record(raw)
            return
        if is_nobat:
            self._nobat_debounce[idx] -= 1
            if self._nobat_debounce[idx] <= 0:
                self._nobat_debounce[idx] = 0
                self._display[idx] = None
            return
        self._nobat_debounce[idx] = 3
        chg_mah = max(raw["charge_mah"], disp["charge_mah"])
        dchg_mah = max(raw["discharge_mah"], disp["discharge_mah"])
        volt = raw["voltage"] if raw["voltage"] > 0 else disp["voltage"]
        self._display[idx] = {
            "voltage": volt,
            "charge_current": raw["charge_current"],
            "discharge_current": raw["discharge_current"],
            "charge_mah": chg_mah,
            "discharge_mah": dchg_mah,
            "status": raw["status"],
            "cycle": raw["cycle"],
            "ck_fail_count": raw.get("ck_fail_count", 0),
            "bridge_online": raw.get("bridge_online", True),
            "executor_online": raw.get("executor_online", True),
            "has_reading": disp["has_reading"] or (raw["voltage"] > 0 or raw["status"] != 0),
        }
    def _reset_display(self):
        self._display = [None] * MODULE_COUNT
        self._nobat_debounce = [0] * MODULE_COUNT

    def _abort_state_as_failed(self, idx, state, reason):
        addr = MODULE_ADDRS[idx]
        self.send_command_reliable(CMD_STOP_CHARGE, addr)
        self.send_command_reliable(CMD_STOP_DISCHARGE, addr)
        if isinstance(state, ModuleTestState):
            self.abort_module_test(idx, state, 'NG')
        else:
            state.completed = True
            self.global_test_results[idx] = {'result': 'NG', 'discharge': getattr(state, 'discharge_max', 0), 'charge2': getattr(state, 'charge_max', 0)}
            self._write_save_line(idx, 'fail_' + str(reason).replace(',', ';'), 0)
            self.on_log_message("ABORT[%#04x] %s" % (addr, reason), "error")
            self.check_loop_all_completed()

    def _check_state_deadline(self, idx, state, now=None):
        if state.completed:
            return False
        if now is None:
            now = time.time()
        no_progress_sec = now - getattr(state, 'last_progress_time', state.stage_start_time)
        stage_sec = now - state.stage_start_time
        if no_progress_sec >= NO_PROGRESS_ABORT_SEC:
            self._abort_state_as_failed(idx, state, "no capacity progress for %.0fs" % no_progress_sec)
            return True
        if stage_sec >= STAGE_HARD_ABORT_SEC:
            self._abort_state_as_failed(idx, state, "stage timeout for %.0fs" % stage_sec)
            return True
        return False

    def _stage_start_command(self, stage):
        if stage == STAGE_CHARGE1 or stage == STAGE_CHARGE2:
            return CMD_START_CHARGE
        if stage == STAGE_DISCHARGE:
            return CMD_START_DISCHARGE
        return None

    def _handle_stale_state(self, idx, state):
        if not self.test_running or state.completed:
            return
        now = time.time()
        addr = MODULE_ADDRS[idx]
        if not hasattr(state, 'last_progress_time'):
            state.last_progress_time = state.stage_start_time
        if not hasattr(state, 'last_retry_time'):
            state.last_retry_time = 0

        last_seen = self.module_last_seen[idx] or state.stage_start_time
        stale_sec = now - last_seen

        if self._check_state_deadline(idx, state, now):
            return

        if stale_sec < DATA_STALE_RETRY_SEC:
            return

        if stale_sec >= DATA_STALE_ABORT_SEC:
            self._abort_state_as_failed(idx, state, "no fresh data for %.0fs" % stale_sec)
            return

        if now - state.last_retry_time < DATA_STALE_RETRY_SEC:
            return

        cmd = self._stage_start_command(state.stage)
        if cmd is None:
            return
        state.last_retry_time = now
        self.on_log_message(
            "STALE[%#04x] no fresh data %.0fs, re-send %s" %
            (addr, stale_sec, CMD_NAMES.get(cmd, "CMD 0x%02X" % cmd)),
            "error")
        self.send_command_reliable(cmd, addr)

    def _state_watchdog(self):
        """Periodic state watchdog: if module status doesn't match expected stage, re-send command."""
        if not self.test_running:
            return
        now = time.time()
        for idx, state in list(self.module_test_states.items()):
            if state.completed:
                continue
            mod = self.modules_data[idx]
            last_seen = self.module_last_seen[idx] or state.stage_start_time
            if mod is None or now - last_seen >= DATA_STALE_RETRY_SEC:
                self._handle_stale_state(idx, state)
                continue
            if not mod.get('executor_online', True):
                self._abort_state_as_failed(idx, state, "executor offline")
                continue
            addr = MODULE_ADDRS[idx]
            status = mod['status']
            elapsed = now - state.stage_start_time
            if self._check_state_deadline(idx, state, now):
                continue

            if (state.stage == STAGE_CHARGE1 or state.stage == STAGE_CHARGE2):
                # Should be charging (0x01) or charge-complete (0x10) or no-battery (0x40)
                if not (status & 0x01) and not (status & 0x10) and not (status & 0x40):
                    if elapsed > 8.0:
                        self.on_log_message("WATCHDOG[%#04x] stage=%s not charging %.0fs, re-send START_CHARGE" % (addr, state.stage, elapsed), "error")
                        if self.serial_thread:
                            status = self.serial_thread.send_command_sync(CMD_START_CHARGE, addr, None, 200)
                            if status == 0:
                                state.stage_start_time = now
                                state.last_retry_time = now
                            else:
                                self.on_log_message("WATCHDOG[%#04x] START_CHARGE ACK failed: %s" % (addr, status), "error")

            elif state.stage == STAGE_DISCHARGE:
                # Should be discharging (0x02) or discharge-complete (0x20) or no-battery (0x40)
                if not (status & 0x02) and not (status & 0x20) and not (status & 0x40):
                    if elapsed > 8.0:
                        self.on_log_message("WATCHDOG[%#04x] stage=DISCHARGE not discharging %.0fs, re-send START_DISCHARGE" % (addr, elapsed), "error")
                        if self.serial_thread:
                            status = self.serial_thread.send_command_sync(CMD_START_DISCHARGE, addr, None, 200)
                            if status == 0:
                                state.stage_start_time = now
                                state.last_retry_time = now
                            else:
                                self.on_log_message("WATCHDOG[%#04x] START_DISCHARGE ACK failed: %s" % (addr, status), "error")

    def update_table(self):
        self.update_module_detail()
        for i in range(MODULE_COUNT):
            mod = self._display[i]
            if mod is None:
                self.main_table.setItem(i, 0, QTableWidgetItem(f"{MODULE_ADDRS[i]}"))
                for c in range(1, self.main_table.columnCount()):
                    self.main_table.setItem(i, c, QTableWidgetItem(""))
                continue
            executor_offline = mod.get('bridge_online') and not mod.get('executor_online', True)
            status_str = ""
            if executor_offline:
                status_str = tr('executor_offline_short', default="Exec Offline")
            elif mod['status'] & 0x01:
                status_str += tr('charge_started_short', default="Charging ")
            if mod['status'] & 0x02:
                status_str += tr('discharge_started_short', default="Discharging ")
            if mod['status'] & 0x04:
                status_str += tr('overcharge_alarm_short', default="Overcharge Alarm ")
            if mod['status'] & 0x08:
                status_str += tr('overdischarge_alarm_short', default="Overdischarge Alarm ")
            if mod['status'] & 0x10:
                status_str += tr('charge_complete_short', default="ChargeComplete ")
            if mod['status'] & 0x20:
                status_str += tr('discharge_complete_short', default="DischargeComplete ")
            if mod['status'] & 0x40:
                status_str = tr('no_battery_short', default="NoBat") + ' '
            if not status_str:
                status_str = tr('idle_status', default="Idle")

            set_cycle = self.set_cycle_nums[i]
            current_cycle = mod['cycle']
            test_status = ""
            if executor_offline:
                test_status = tr('executor_offline', default="Bridge OK / Executor Offline")
            elif mod['status'] & 0x40:
                test_status = tr('no_battery', default="No Battery")
            elif set_cycle > 0:
                if current_cycle >= set_cycle:
                    test_status = tr('completed_status', current_cycle, set_cycle)
                else:
                    if mod['status'] & 0x01:
                        test_status = tr('charging_status', current_cycle+1, set_cycle)
                    elif mod['status'] & 0x02:
                        test_status = tr('discharging_status', current_cycle+1, set_cycle)
                    else:
                        test_status = tr('idle_cycle_status', current_cycle, set_cycle)
            else:
                test_status = tr('cycle_not_set')

            if self.loop_test_active and i in self.module_test_states:
                state = self.module_test_states[i]
                if isinstance(state, LoopModuleState):
                    test_status = tr('loop_status', state.current_cycle, state.total_cycles)
                else:
                    test_status = "Loop: -"

            alarm_type = ""
            if mod['status'] & 0x04:
                alarm_type = tr('overcharge_alarm_short')
            elif mod['status'] & 0x08:
                alarm_type = tr('overdischarge_alarm_short')

            qc_result = ""
            test_caps = ""
            row_color = QColor(255, 255, 255)
            if executor_offline:
                row_color = QColor(255, 245, 180)
                alarm_type = tr('executor_offline_short', default="Exec Offline")

            if self.global_test_results[i]:
                result = self.global_test_results[i]
                qc_result = result['result']
                if qc_result == "OK":
                    row_color = QColor(200, 255, 200)
                elif qc_result == "NG":
                    row_color = QColor(255, 200, 200)
                    if alarm_type:
                        alarm_type += " + Capacity NG"
                    else:
                        alarm_type = "Capacity NG"
                elif qc_result == "NO_BAT":
                    qc_result = "NG"
                    row_color = QColor(220, 220, 220)
                discharge = result.get('discharge', 0)
                charge2 = result.get('charge2', 0)
                test_caps = f"D:{discharge} C:{charge2}"

            if (mod['status'] & 0x04) or (mod['status'] & 0x08):
                row_color = QColor(255, 200, 200)

            self.main_table.setItem(i, 0, QTableWidgetItem(f"{MODULE_ADDRS[i]}"))
            self.main_table.setItem(i, 1, QTableWidgetItem(str(mod['voltage'])))
            self.main_table.setItem(i, 2, QTableWidgetItem(str(mod['charge_current'])))
            self.main_table.setItem(i, 3, QTableWidgetItem(str(mod['discharge_current'])))
            self.main_table.setItem(i, 4, QTableWidgetItem(str(mod['charge_mah'])))
            self.main_table.setItem(i, 5, QTableWidgetItem(str(mod['discharge_mah'])))
            self.main_table.setItem(i, 6, QTableWidgetItem(status_str))
            self.main_table.setItem(i, 7, QTableWidgetItem(str(mod['cycle'])))
            self.main_table.setItem(i, 8, QTableWidgetItem(test_status))
            self.main_table.setItem(i, 9, QTableWidgetItem(qc_result))
            self.main_table.setItem(i, 10, QTableWidgetItem(test_caps))
            self.main_table.setItem(i, 11, QTableWidgetItem(alarm_type))
            self.main_table.setItem(i, 12, QTableWidgetItem(str(mod.get('ck_fail_count', 0))))

            for col in range(13):
                item = self.main_table.item(i, col)
                if item:
                    item.setBackground(row_color)

        self.main_table.resizeColumnsToContents()

    def _safe_set_cell(self, row, col, text, color=None):
        existing = self.history_table.item(row, col)
        if existing and existing.text():
            try:
                old_val = int(existing.text())
                new_val = int(text)
                if new_val < old_val and (color is not None or new_val == 0):
                    self.on_log_message('[GUARD] skip R%dC%d: %d < %d' % (row+1, col+1, new_val, old_val), "info")
                    return
            except ValueError:
                pass
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)
        if color:
            item.setForeground(color)
        self.history_table.setItem(row, col, item)

    def update_history_table(self):
        for i in range(MODULE_COUNT):
            records = self.cycle_records[i]
            mod = self.modules_data[i]
            state = self.module_test_states.get(i)
            active = mod and state and not state.completed
            max_cols = self.history_table.columnCount()
            # only clear cells that should be empty:
            # - if module has no data at all, clear entire row
            # - otherwise clear cells beyond committed records + 2 (reserve one extra cycle for live overlay)
            if not records and not active:
                for c in range(max_cols):
                    self.history_table.setItem(i, c, QTableWidgetItem(""))
                continue
            keep_cols = len(records) * 2 + 2  # committed + 1 cycle live reserve
            for c in range(keep_cols, max_cols):
                self.history_table.setItem(i, c, QTableWidgetItem(""))
            # always show committed records (with cutoff voltage)
            for j, rec in enumerate(records):
                for key, offset in [("charge", 0), ("discharge", 1)]:
                    val = rec.get(key)
                    if val is not None:
                        col = j * 2 + offset
                        if col < max_cols:
                            vkey = "chg_volt" if key == "charge" else "dchg_volt"
                            volt = rec.get(vkey)
                            txt = str(val)
                            if volt is not None:
                                txt += "\n(V: %.2fV)" % (volt / 1000.0)
                            self._safe_set_cell(i, col, txt)
            # real-time live mAh overlay (blue) for active modules
            if active:
                if self.loop_test_active and isinstance(state, LoopModuleState):
                    if state.stage == STAGE_CHARGE1:
                        col = len(records) * 2
                        val = max(0, state.charge_max - state.charge_start)
                    elif state.stage == STAGE_DISCHARGE:
                        col = max(0, len(records) - 1) * 2 + 1
                        val = max(0, state.discharge_max - state.discharge_start)
                    else:
                        continue
                    if col < max_cols and val is not None:
                        self._safe_set_cell(i, col, str(val), QColor(0, 100, 200))
                elif self.test_running and not self.loop_test_active and isinstance(state, ModuleTestState):
                    if state.stage == STAGE_CHARGE1:
                        col = 0
                        val = max(0, mod['charge_mah'] - state.charge1_start)
                    elif state.stage == STAGE_DISCHARGE:
                        col = 1
                        val = max(0, state.discharge_max - state.discharge_start)
                    elif state.stage == STAGE_CHARGE2:
                        col = 2
                        val = max(0, mod['charge_mah'] - state.charge2_start)
                    else:
                        continue
                    if col < max_cols and val is not None:
                        self._safe_set_cell(i, col, str(val), QColor(0, 100, 200))
    
    def on_alarm(self, message):
        self.on_log_message(message, "info")
        if self.test_running:
            if self.serial_thread:
                self.serial_thread.send_all_command(CMD_ALL_BUZZER_ON)
                self.buzzer_timer.start(30000)
            return
        now = time.time()
        if now - self.last_alarm_time > 60:
            self.last_alarm_time = now
            QMessageBox.warning(self, tr('warning'), message)

    def on_serial_error(self, err):
        QMessageBox.critical(self, tr('error'), err)

    def on_single_command(self, cmd):
        idx = self.module_combo.currentIndex()
        addr = MODULE_ADDRS[idx]
        if self.serial_thread:
            self.serial_thread.send_command(cmd, addr)

    def on_log_message(self, msg, category="info"):
        if self.log_window:
            self.log_window.append_log(msg)
        else:
            # 也可输出到状态栏或控制台
            pass

    def closeEvent(self, event):
        self._stop_auto_save()
        self.watchdog_timer.stop()
        self.test_stage_timer.stop()
        self.buzzer_timer.stop()
        self.current_check_timer.stop()
        if not self.test_running:
            self.history_timer.start(500)
        if self.serial_thread:
            self.serial_thread.stop()
        if self.log_window:
            self.log_window.close()
        event.accept()


if __name__ == "__main__":
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
    sys.exit(app.exec_())
